import os
import re
import time
import shutil
import tempfile
import sys

import openpyxl
from openpyxl import Workbook
import win32com.client as win32
from pathlib import Path

class AnalyzeWB:
    def __init__(self, **kwargs):
        self.error_sh=kwargs['error_sh']
        self.error_row=kwargs['error_row']
        self.xl=kwargs['xl']
        # print('analyzewb 1')
        try:
            # print('analyzewb 2')
            self.wb=self.xl.Workbooks.Open(kwargs['wb'])
            # print('analyzewb 3')
        except:
            try:
                if Path(Path(tempfile.gettempdir()) / 'gen_py').exists():
                    shutil.rmtree(Path(Path(tempfile.gettempdir()) / 'gen_py'))
                    print('I deleted gen_py :)')
                self.wb=self.xl.Workbooks.Open(kwargs['wb'])
            except:
                print('Error when trying to open facture, failed to close gen_py or some issue with: ', kwargs['wb'])
                
                self.error_sh.cell(row=self.error_row, column=3, value='Error, couldn\'t open workbook.')
                sys.exit(0)

        # try:
        #     # print('analyzewb 5')
        #     self.shfacture=self.wb.Sheets('Facture')
        #     # print('analyzewb 6')
        #     # return True
        # except:
        #     self.error_sh.cell(row=self.error_row, column=4, value='Error, No Facture sheet.')
        #     pass
        #     # return False
        
    def has_facture_sh(self):
        try:
            self.shfacture=self.wb.Sheets('Facture')
            return True
        except:
            self.error_sh.cell(row=self.error_row, column=4, value='Error, No Facture sheet.')
            return False

    def has_annexe_sh(self):
        try:
            self.shannexe=self.wb.Sheets('Annexe')
            if not self.shannexe.Visible == win32.constants.xlSheetHidden:
                return True
            else:
                raise Exception
        except:
            try:
                self.shannexe=self.wb.Sheets('Annexe(2)')
                if not self.shannexe.Visible == win32.constants.xlSheetHidden:
                    return True
                else:
                    raise Exception
            except:
                try:
                    self.shannexe=self.wb.Sheets('Annexe (2)')
                    if not self.shannexe.Visible == win32.constants.xlSheetHidden:
                        return True
                    else:
                        raise Exception
                except:
                    self.error_sh.cell(row=self.error_row, column=5, value='Error, No Annexe sheet.')
                    return False

    def is_proper_workbook(self):
        try:
            if self.wb:
                return True
        except:
            self.error_sh.cell(row=self.error_row, column=6, value='Error, Workbook is not proper facture.')
            return False

    def is_facture_wb(self):
        if self.has_facture_sh() and self.has_annexe_sh():
            return True
        else:
            return False

    def get_wb(self):
        return self.wb

    def get_facture_sheet(self):
        if self.has_facture_sh():
            return self.shfacture
        else:
            return False
        # try:
        #     self.shfacture=self.wb.Sheets('Facture')
        #     return self.shfacture
        # except:
        #     return False

    def get_annexe_sheet(self):
        if self.has_annexe_sh():
            return self.shannexe
        else:
            return False
        # try:
        #     self.shannexe=self.wb.Sheets('Annexe')
        #     return self.shannexe
        # except:
        #     try:
        #         self.shannexe=self.wb.Sheets('Annexe(2)')
        #         return self.shannexe
        #     except:
        #         return False

    def bool_test(self):
        try:
            if self.wb:
                return True
        except:
            return False


class AnalyzeFactureSheet:
    def __init__(self, *args, **kwargs):
        self.shfacture=kwargs['sh']
        self.error_sh=kwargs['error_sh']
        self.error_row=kwargs['error_row']
        # self.xl=kwargs['xl']

        self.leftcol='A'
        self.leftcolno=1
        self.rightcol='F'
        self.rightcolno=6
        self.bottomrowno=60
        self.DATA_FOR_BOTTOM_SEARCH=['ORIGINE', 'МЕСТО ИЗГОТОВЛЕНИЯ', 'ÖNDÜRILEN ÝERI', 'МЕСТО ИЗГОТОВЛЕНИЯ', 'PROVENANCE', 'GELÝÄN  ÝERI']
        self.DATA_FOR_LAST_COL_SEARCH=['VALEUR', 'ОБЩАЯ ЦЕНА', 'UMUMY BAHASY']
        self.DATA_LAST_COL_TO_LOOP = [5, 6, 7]

        # Determining last row
        # Loop through columns
        self.temp_bottom_row_found=False
        
        for j in range(self.bottomrowno, self.bottomrowno + 15):
            # print('Inside row no: ', j)
            if self.temp_bottom_row_found:
                break
            for i in range(1, self.rightcolno+1):
                # print('In column: ', i)
                if len(str(self.shfacture.Cells(j, i).Value)) == 0 or str(self.shfacture.Cells(j, i).Value) == 'None':
                    self.temp_bottom_row_found=True
                    self.bottomrowno=j
                    # print('bottom row changed: ', self.bottomrowno)
                    # break
                else:
                    # print('length if not zero: ', str(self.shfacture.Cells(j, i).Value))
                    self.temp_bottom_row_found=False
                    break

        # Determining last column
        self.last_col_match_found=False
        self.temp_rightcolno=0
        self.temp_rightcolletter=''
        # self.try_count=0
        # First loop only column F
        for i in range(1, self.bottomrowno+1):
                if self.last_col_match_found:
                    break
                # Column no of F is 6
                self.temp_cell_value=str(self.shfacture.Cells(i, 6).Value).lower()

                for xval in self.DATA_FOR_LAST_COL_SEARCH:
                    self.xval1=str(xval).lower()
                    if self.temp_cell_value.__contains__(self.xval1):
                        self.last_col_match_found=True
                        self.temp_rightcolno=6
                        self.temp_rightcolletter='F'
                        # print(self.temp_rightcolletter)
                        break

        # Loop trhough 3 columns
        # Only if column F does not match our criteria
        if not self.last_col_match_found:
            for j in self.DATA_LAST_COL_TO_LOOP:
                if self.last_col_match_found:
                    break
                # Loop through all rows of specific column
                for i in range(1, self.bottomrowno+1):
                    if self.last_col_match_found:
                        break
                    self.xcell=str(self.shfacture.Cells(i, j).Value).lower()
                    
                    
                    for xval in self.DATA_FOR_LAST_COL_SEARCH:
                        self.xval1=str(xval).lower()
                        if self.xcell.__contains__(self.xval1):
                            self.last_col_match_found=True
                            self.temp_rightcolno=j
                            self.temp_rightcolletter=str(self.shfacture.Cells(i, j).Address)[1:2]
                            # print(self.temp_rightcolletter)
                            break
        

        if self.temp_rightcolno!=0:
            self.rightcolno=self.temp_rightcolno
            self.rightcol=self.temp_rightcolletter

            
        # Organize page break
        # self.leftcol='A'
        # self.leftcolno=1
        # self.rightcol='F'
        # self.rightcolno=6
        # self.bottomrowno=60

        self.print_area=self.leftcol + '1:' + self.rightcol + str(self.bottomrowno)
        self.shfacture.PageSetup.Zoom = False
        self.shfacture.PageSetup.FitToPagesTall = 1
        self.shfacture.PageSetup.FitToPagesWide = 1
        self.shfacture.PageSetup.PrintArea = self.print_area


    def get_facture_sheet(self):
        return self.shfacture

    def get_sheet_name(self):
        return self.shfacture.Name


class AnalyzeAnnexeSheet:
    def __init__(self, *args, **kwargs):
        self.shannexe=kwargs['sh']
        self.error_sh=kwargs['error_sh']
        self.error_row=kwargs['error_row']
        # self.xl=kwargs['xl']

        self.leftcol='A'
        self.leftcolno=1
        self.rightcol='V'
        self.rightcolno=22
        self.descriptioncol='N'
        self.descriptioncolno=14
        self.bottomrowno=self.shannexe.Cells(self.shannexe.Rows.Count, self.rightcolno).End(win32.constants.xlUp).Row
        # print('Annexe last row: ', self.bottomrowno)
        self.DATA_FOR_BOTTOM_SEARCH=['Total']
        self.DATA_FOR_LAST_COL_SEARCH=['Prix Total']
        self.DATA_LAST_COL_TO_LOOP = [20, 21, 22, 23, 24, 25]
        self.DATA_DESC_TO_LOOP = ['Description', 'Наименование', 'Atlandyrylyşy']

        # Determining last column
        self.last_col_match_found=False
        self.temp_rightcolno=0
        self.temp_rightcolletter=''
        # First loop only column V
        for i in range(1, self.bottomrowno+1):
                if self.last_col_match_found:
                    break
                # Column no of V is 22
                self.temp_cell_value=str(self.shannexe.Cells(i, 22).Value).lower()

                for xval in self.DATA_FOR_LAST_COL_SEARCH:
                    self.xval1=str(xval).lower()
                    if self.temp_cell_value.__contains__(self.xval1):
                        self.last_col_match_found=True
                        self.temp_rightcolno=22
                        self.temp_rightcolletter='V'
                        # print(self.temp_rightcolletter)
                        break

        # Loop trhough 5 columns
        # Only if column V does not match our criteria
        if not self.last_col_match_found:
            for j in self.DATA_LAST_COL_TO_LOOP:
                if self.last_col_match_found:
                    break
                # Loop through all rows of specific column
                for i in range(1, self.bottomrowno+1):
                    if self.last_col_match_found:
                        break
                    self.xcell=str(self.shannexe.Cells(i, j).Value).lower()
                    for xval in self.DATA_FOR_LAST_COL_SEARCH:
                        self.xval1=str(xval).lower()
                        if self.xcell.__contains__(self.xval1):
                            self.last_col_match_found=True
                            self.temp_rightcolno=j
                            self.temp_rightcolletter=str(self.shannexe.Cells(i, j).Address)[1:2]
                            # print(self.temp_rightcolletter)
                            break
    
        if self.temp_rightcolno!=0:
            self.rightcolno=self.temp_rightcolno
            self.rightcol=self.temp_rightcolletter
            self.bottomrowno=self.shannexe.Cells(self.shannexe.Rows.Count, self.rightcolno).End(win32.constants.xlUp).Row
            # print('Annexe last row: ', self.bottomrowno)


        # Determining last row
        # Loop through columns
        self.temp_bottom_row_found=False
        # self.temp_bottom_total_row=self.bottomrowno
        # self.total=self.shannexe.Cells(self.bottomrowno, self.rightcolno).Value
        # self.total_row_color=self.shannexe.Cells(self.bottomrowno, self.rightcolno).Interior.Color
        # self.description_is_in_n=False
        self.temp_total=''
        self.temp_last_total_row=self.shannexe.Cells(self.shannexe.Rows.Count, self.descriptioncolno).End(win32.constants.xlUp).Row
        if not self.bottomrowno==self.temp_last_total_row:
            print('Bottom row number and total row number are no same, should be same')
            self.error_sh.cell(row=self.error_row, column=7, value='Bottom row number and total row number are no same, should be same')

        # First check last column, if it contains a total, and assign total sum to a value and convert it to fload
        if str(self.shannexe.Cells(self.temp_last_total_row, self.descriptioncolno).Value).lower().__contains__('total'.lower()):
            try:
                self.temp_total=float(self.shannexe.Cells(self.temp_last_total_row, self.rightcolno).Value)
                # print('This is very last total and it is a float(temp_total): ', self.temp_total)
            except:
                # print('Last row\'s total value is not a float')
                # print(self.shannexe.Cells(self.temp_last_total_row, self.descriptioncolno).Value)
                # print(self.shannexe.Cells(self.temp_last_total_row, self.rightcolno).Value)
                
                self.error_sh.cell(row=self.error_row, column=8, value='Last row\'s total value is not a float')
                pass
        
        # Loop through upper rows to find nearest total and compare two totals, if they are same then assign new total as general total
        # print('temp_lat_total_row to be reversed: ', self.temp_last_total_row)
        for xrow in reversed(range(1, self.temp_last_total_row)):
            # print('xrow: ', xrow)
            # print('description on: ', self.descriptioncolno)
            self.xcell_value=self.shannexe.Cells(xrow, self.descriptioncolno).Value
            # print(self.xcell_value)
            if str(self.xcell_value).lower().__contains__('total') and not str(self.xcell_value).lower().__contains__('page'):
                try:
                    self.new_total_value=float(self.shannexe.Cells(xrow, self.rightcolno).Value)
                    if self.new_total_value==self.temp_total:
                        self.temp_last_total_row=xrow
                    else:
                        break
                except:
                    # print('NEW possible total\'s total value is not a float')
                    # print(self.shannexe.Cells(xrow, self.descriptioncolno).Value)
                    # print(self.shannexe.Cells(xrow, self.rightcolno).Value)
                    self.error_sh.cell(row=self.error_row, column=9, value='NEW possible total\'s total value is not a float')

        # Check temp_last_row, if it is bigger than 50 then use it as bottomrow, else use initial value of bottomrow
        if self.temp_last_total_row>50:
            self.bottomrowno=self.temp_last_total_row



        # # Decide which column is description column, which contains the word 'total'
            # # Normally it should be in column N, first check N, (14) if it is not there
            # # Then Loop through all columns to find the column that contains the word 'total'
            # # Loop through columns N and find words Description, Наименование, Atlandyrylyşy
            # for words in self.DATA_DESC_TO_LOOP:
            #     if not self.description_is_in_n:
            #         for xrow in range(1,10):
                        
            #             if str(self.shannexe.Cells(xrow, self.descriptioncolno).Value).lower().__contains__(words.lower()):
            #                 self.description_is_in_n=True
            #                 break
                    
                        
                    
            # self.total_word_value=str(self.shannexe.Cells(self.bottomrowno, self.descriptioncolno).Value).lower()
            # if not self.total_word_value.__contains__('total'):
            #     if not self.temp_bottom_row_found:
            #         for colno in range(1, self.rightcolno+1):
            #             if str(self.shannexe.Cells(self.bottomrowno, colno).Value).lower().__contains__('total'):
            #                 self.descriptioncolno=colno
            #                 self.temp_bottom_row_found=True
            #                 break

            #     pass



            # # For each cell in row
            # if not self.description_is_in_n:
            #     for j in reversed(range(self.bottomrowno)):
            #         # print('Inside row no: ', j)
                    # if self.temp_bottom_row_found:
                    #     break


                    # First, check only for columns N, (14)

                    # For each column
                    # for i in range(1, self.rightcolno+1):
                    #     # print('In column: ', i)
                    #     if len(str(self.shannexe.Cells(j, i).Value)) == 0 or str(self.shannexe.Cells(j, i).Value) == 'None':
                    #         self.temp_bottom_row_found=True
                    #         self.bottomrowno=j
                    #         # print('bottom row changed: ', self.bottomrowno)
                    #         # break
                    #     else:
                    #         # print('length if not zero: ', str(self.shfacture.Cells(j, i).Value))
                    #         self.temp_bottom_row_found=False
                    #         break

        # Organize page break
        self.print_area=self.leftcol + '1:' + self.rightcol + str(self.bottomrowno)
        self.shannexe.PageSetup.Zoom = False
        self.shannexe.PageSetup.FitToPagesTall = False
        self.shannexe.PageSetup.FitToPagesWide = 1
        self.shannexe.PageSetup.PrintArea = self.print_area
        # print('print area: ', self.print_area)

    def get_sheet_name(self):
        return self.shannexe.Name


# if __name__=='__main__':
    #     try:
    #         start=time.perf_counter()
    #         xlCalculationManual = -4135
    #         xlCalculationAutomatic = -4105

    #         testpath=Path(r'D:\BYTK_Facturation\7. MT\xxx')
    #         xl=win32.gencache.EnsureDispatch('Excel.Application')
            
    #         xl.ScreenUpdating = False
    #         xl.EnableEvents = False
    #         xl.DisplayAlerts = False
    #         xl.AskToUpdateLinks = False
            

    #         xl.Visible=True

    #         error_wb=Workbook()
    #         error_sh=error_wb.active
    #         # xl.Calculation = xlCalculationManual

    #         for i, wb in enumerate(testpath.glob('*.xls*'), start=1):
    #             wb_analyzer=AnalyzeWB(wb=wb, error_sh=error_sh, error_row=i)
    #             error_sh.cell(row=i, column=1, value=str(wb.stem))
                
    #             # print(i, ' : ', wb.name)
    #             # print(i, ' : ', wb_analyzer.is_proper_workbook())
    #             # print(i, ' : ', wb_analyzer.is_facture_wb())
    #             if  wb_analyzer.is_proper_workbook() and wb_analyzer.is_facture_wb():
    #                 wb_current=wb_analyzer.get_wb()
    #                 shfacture=wb_analyzer.get_facture_sheet()
    #                 shannexe=wb_analyzer.get_annexe_sheet()

    #                 wb_current.Worksheets([shfacture.Name, shannexe.Name]).Select()

    #                 # Analyze Facture Sheet
    #                 shfacture_analyze=AnalyzeFactureSheet(sh=shfacture, error_sh=error_sh, error_row=i)
    #                 # print(shfacture_analyze.get_sheet_name(), ' right col no: ', shfacture_analyze.rightcolno)
    #                 # print(shfacture_analyze.get_sheet_name(), ' bottom row no: ', shfacture_analyze.bottomrowno)
                    
    #                 # Analyze Annexe Sheet
    #                 shannexe_analyze=AnalyzeAnnexeSheet(sh=shannexe, error_sh=error_sh, error_row=i)
    #                 # print(shannexe_analyze.get_sheet_name(), ' right col no: ', shannexe_analyze.rightcolno)
    #                 # print(shannexe_analyze.get_sheet_name(), ' bottom row no: ', shannexe_analyze.bottomrowno)

    #                 filename=str(wb.stem)+ '.pdf'
    #                 dest_full_name= r'' + str(testpath / filename) + ''
    #                 try:
    #                     xl.ActiveSheet.ExportAsFixedFormat(0, dest_full_name)
    #                     error_sh.cell(row=i, column=2, value=str(wb.stem) + ' converted successfully :)')
    #                 except:
    #                     print('Couldn\'t convert to PDF. It looks like PDF with same name is open, \n Please close and try again.')
    #                     error_sh.cell(row=i, column=2, value=str(wb.stem) + ' Error when converting :(')

    #                 # Close workbooks as soon pdf is created.
    #                 wb_current.Saved=True
    #                 wb_current.Close()

    #     except:
    #         xl.ScreenUpdating = True
    #         xl.EnableEvents = True
    #         xl.DisplayAlerts = True
    #         xl.AskToUpdateLinks = True
    #         # xl.Calculation = xlCalculationAutomatic

    #         xl.Quit()
    #     finally:
    #         xl.ScreenUpdating = True
    #         xl.EnableEvents = True
    #         xl.DisplayAlerts = True
    #         xl.AskToUpdateLinks = True
    #         # xl.Calculation = xlCalculationAutomatic

    #         xl.Quit()
        
    #     error_wb.save(Path.cwd() / 'Here are the errors.xlsx')
    #     finish=time.perf_counter()
    #     print(f'Total time spent: {round(finish-start, 2)}')
        
    """
    Class openpyxl

    FACTURE
    sheet Facture
            # if there is no sheet Facture then skip next file, and copy this file into error folder
            # if there is Facture sheet then check for some criterias
        # # left column A
        # # right column F
        #     check if F column contains
        #     1 - TK and some number
        #     2 - VALEUR
        #     3 - ОБЩАЯ ЦЕНА
        #     4 - UMUMY BAHASY
        #     if not then search for G column,
        #     then keep going right



        # page break
        # fit to one page


        # top row 1
        # bottom row 60
        #     check if this row is empty, if not skip one more below and do next checks
        #     check if one of upper rows (any) contains 
        #     ORIGINE
        #     МЕСТО ИЗГОТОВЛЕНИЯ 
        #     ÖNDÜRILEN ÝERI 
        #     PROVENANCE
        #     ПРОИСХОЖДЕНИЕ
        #     GELÝÄN  ÝERI
        #     if you find match of these, select next empty row as last row


    ANNEXE
    # sheet Annexe
        # If there is no Annexe sheet then skip to next file and copy this file into error folder

        page break
        fit columns into one page

        left column A
        top row 1

        right column V
            check if this column contains on of the followings
                Prix Total
                Şertnamanyň umumy bahasy
                if does not contain then
                    check all columns starting from T to X
                    if one of them contains any of criteria assign that to be right column

            
        description column is N
            check if N contains one of the following
            -Designations
            -Наименование
            -Atlandyrylyşy
            if does not contain then check first 10 rows of every row starting
                from 1 to 30, if any match any criteria assign that one as 
                description column


        last row
            find the last row that contains value from right columns.
                check if it contains 'total' in description column.
                if its row is less than 100 then assign that as last row

                else if it is bigger than 100 then do the following
                    loop throug description column to upper, 
                    if there is any value and that value does not contain total word then assign previous total as last row.
                        else if there is a value that contains 'total' and its background is same with prevous last row
                        and if its last column sum is equal to prevous total row sum, then it means 
                        that last page is empty, 
                        it means that we assign newly found total row as last row,
                            and we loop through upper until we find material, which is the cell with a any value that does not contains 'total'


        page breaks
            for vertical page break
                add vpagebreak manually, and check that there is only one vpagebreak
            for horizontal page break
                check all rows, and be sure that every page break is only after total page, if not do necessary steps to make it so.


    """
