import time
import os
import re
from multiprocessing import Pool
import shutil
import gc 

import psutil
from tkinter import *
from tkinter import ttk
from pathlib import Path
import win32com.client as win32
import openpyxl
from openpyxl import Workbook

from xl_to_pdf_classes import AnalyzeWB, AnalyzeFactureSheet, AnalyzeAnnexeSheet

class TabXltopdf:
    def __init__(self, *args, **kwargs):
        # Fetch user's previous selections
        self.users_facture_path = ''
        self.users_result_pdf_path = ''
        try:
            with open(Path('users_selections') / 'facture_path.txt', 'r') as f:
                self.users_facture_path = f.readline()
            with open(Path('users_selections') / 'result_pdf_facture.txt', 'r') as f:
                self.users_result_pdf_path = f.readline()
        except:
            pass

        self.tab = kwargs['tab']

        # Address Selection selection
        self.source_value=StringVar()
        self.destination_value=StringVar()

        # Selection of Excel verions of invoice
        self.excel_invoices = ttk.Entry(
            self.tab, width=50, font='Helvetica 12', textvariable=self.source_value)
        self.excel_invoices.insert(INSERT, Path(self.users_facture_path))
        self.excel_invoices.grid(
            row=1, column=1, sticky=W, padx=60, pady=20, ipadx=20, ipady=10)
        self.browse_excel_invoices = ttk.Button(
            self.tab, width=20, text="Excel verion of Invoices you wish to convert.", style='XltopdfBrowse.TButton')
        self.browse_excel_invoices.grid(
            row=1, column=2, columnspan=2, padx=5, ipadx=10, ipady=5)

        # Result PDF Folder selection
        self.result_folder = ttk.Entry(self.tab, width=50, font='Helvetica 12', textvariable=self.destination_value)
        self.result_folder.insert(INSERT, Path(self.users_result_pdf_path))
        self.result_folder.grid(row=2, column=1, sticky=W,
                                padx=60, pady=20, ipadx=20, ipady=10)
        self.browse_result_folder = ttk.Button(
            self.tab, width=20, text="Where do you want to save converted PDF invoices?", style='XltopdfBrowse.TButton')
        self.browse_result_folder.grid(
            row=2, column=2, columnspan=2, padx=5, ipadx=10, ipady=5)

        # RadioButtons
        self.rbtn_value = IntVar(value=1)
        self.onlynew_rbtn = ttk.Radiobutton(
            self.tab, text='Convert only new factures', variable=self.rbtn_value, value=1, style='Xltopdf.TRadiobutton', width = 40)
        self.replace_rbtn = ttk.Radiobutton(self.tab, text='Convert all from beginning and replace existing PDFs (Takes time)',
                                            variable=self.rbtn_value, value=0, style='Xltopdf.TRadiobutton', width = 40)
        self.onlynew_rbtn.grid(row=3, column=1, sticky=W,
                               padx=60, pady=20, ipadx=20, ipady=10)
        self.replace_rbtn.grid(row=4, column=1, sticky=W,
                               padx=60, pady=20, ipadx=20, ipady=10)

        # Convert Button
        # self.converter=ConvertFiles()

        self.convert_button = ttk.Button(
            self.tab, text='Convert', style='XltopdfConvertButton.TButton')
        self.convert_button.grid(row=int(self.replace_rbtn.grid_info()[
                                 'row'])+1, column=2, columnspan=2, sticky=E, padx=10, pady=10, ipadx=50, ipady=20)
        self.convert_button.config(command=self.convert_files)

    def convert_files(self):
        # Save user's selections to a file
        with open(Path('users_selections') / 'facture_path.txt', 'w') as f:
            f.write(self.source_value.get())
        with open(Path('users_selections') / 'result_pdf_facture.txt', 'w') as f:
            f.write(self.destination_value.get())

        try: 
            self.start=time.perf_counter()
            print('time started')
            xl=win32.gencache.EnsureDispatch('Excel.Application')
            print('xl started')
            xl.ScreenUpdating = False
            xl.EnableEvents = False
            xl.DisplayAlerts = False
            xl.AskToUpdateLinks = False
            
            xl.Visible=False

            print('before source declaration')
            self.source=Path(self.source_value.get())
            self.destination=Path(self.destination_value.get())
            # If our suggested destination path does not exist, then create
            if not self.destination.exists():
                self.destination.mkdir(parents=True, exist_ok=True)

            self.error_destination=self.destination / 'errors'
            if not self.error_destination.exists():
                self.error_destination.mkdir(parents=True, exist_ok=True)

            self.selection=self.rbtn_value.get()
            # print('1')

            self.error_wb=Workbook()
            self.error_sh=self.error_wb.active

            # print('2')
            count_converted = 0
            for i, wb in enumerate(self.source.glob('*.xls*'), start=1):
                count_converted += 1
                print('\n' + f'Working with number {count_converted}: '.ljust(25, ' '), wb.name)
                if '~$' in str(wb.stem):
                    continue      
                # First Check user's replace or not selection, handle the request upon.
                self.filename=str(wb.stem)+ '.pdf'
                self.dest_full_name= r'' + str(self.destination / self.filename) + ''
                
                # print('2.1')
                # If user does not want to replace
                # Keep old files and skip current facture
                if self.rbtn_value.get()==1: # Wants to keep old, does not want to replace
                    if Path(self.dest_full_name).exists():
                        print('Already exists, skipping: '.ljust(25, ' '), self.filename)
                        continue
                # print('2.2 after replace check')
                
                try:
                    self.wb_analyzer=AnalyzeWB(wb=wb, error_sh=self.error_sh, error_row=i, xl=xl)
                    # print('3')
                except:
                    try:
                        self.error_file_name='error_' + str(wb.name)
                        self.error_full_source=self.source / str(wb.name)
                        self.error_full_destination=self.error_destination / self.error_file_name
                        shutil.copy(str(self.error_full_source), str(self.error_full_destination))

                        # Update error logging excel
                        self.error_sh.cell(row=i, column=2, value=self.error_file_name)
                        continue
                    except:
                        continue
                    
                # print('3.1')
                self.error_sh.cell(row=i, column=1, value=str(wb.stem))
                # print('4')
                if  self.wb_analyzer.is_proper_workbook() and self.wb_analyzer.is_facture_wb():
                    # print('6')

                    self.wb_current=self.wb_analyzer.get_wb()
                    self.shfacture=self.wb_analyzer.get_facture_sheet()
                    self.shannexe=self.wb_analyzer.get_annexe_sheet()

                    # print('8')
                    # Select both sheets (Facture and Annexe)
                    # print(self.shfacture.Name, self.shannexe.Name)
                    self.wb_current.Worksheets([self.shfacture.Name, self.shannexe.Name]).Select()
                    # print('8.1')
                    # Analyze Facture Sheet
                    self.shfacture_analyze=AnalyzeFactureSheet(sh=self.shfacture, error_sh=self.error_sh, error_row=i)
                    # Analyze Annexe Sheet
                    self.shannexe_analyze=AnalyzeAnnexeSheet(sh=self.shannexe, error_sh=self.error_sh, error_row=i)
                    # print('9')
                    # Export to PDF
                    try:
                        # print('destination address', self.dest_full_name)
                        xl.ActiveSheet.ExportAsFixedFormat(0, self.dest_full_name)
                        # print('10')
                        self.error_sh.cell(row=i, column=2, value=str(wb.stem) + ' converted successfully :)')
                        print('Converted successfully: '.ljust(25, ' ') + str(wb.stem))
                    except:
                        print('Couldn\'t convert to PDF. It looks like PDF with same name is open, \n Please close and try again.')
                        self.error_sh.cell(row=i, column=2, value=str(wb.stem) + ' Error when converting :(')
                    finally:
                        # Close excel without saving
                        # print('Closing xl: ', self.wb_analyzer.wb.Name)
                        self.wb_analyzer.wb.Close(False)
                        # print('Closed successfully.')
                        if (i % 10) == 0:
                            gc.collect()

                    # Close workbooks as soon pdf is created.
                    try:
                        self.wb_current.Saved=True
                        self.wb_current.Close()
                    except:
                        pass
                else: # if workbook is not a proper, which is not a facture excel
                    print('It is not a facture excel: ', str(wb.name))
                    self.error_file_name='error_' + str(wb.name)
                    self.error_full_source=self.source / str(wb.name)
                    self.error_full_destination=self.error_destination / self.error_file_name
                    shutil.copy(str(self.error_full_source), str(self.error_full_destination))
                    # pass 

        except Exception as e:
            print('Exception in one thread: ', e)

            # xl.Calculation = xlCalculationAutomatic

            try:
                print('Exception again in one thread: ', e)

                self.wb_current.Saved=True
                self.wb_current.Close()
            except:
                pass 

            # xl.Quit()

        finally:
            # xl.ScreenUpdating = True
            # xl.EnableEvents = True
            # xl.DisplayAlerts = True
            # xl.AskToUpdateLinks = True
            # xl.Calculation = xlCalculationAutomatic

            # xl.Quit()
            pass
        
        self.error_wb.save(Path.cwd() / 'Here are the errors.xlsx')
        self.finish=time.perf_counter()
        print(f'\nTotal time spent for PDF convertion: {round(self.finish-self.start, 2)}')


# Working without Pool
    # class ConvertFiles:
    #     def __init__(self):
    #         self.wb_current=''
    #         pass 

    #     def convert_files(self, **kwargs):
    #         try:
    #             self.start=time.perf_counter()

    #             self.source=kwargs['source']
    #             self.destination=kwargs['destination']
    #             self.selection=kwargs['selection']
    #             print('1')

    #             self.error_wb=Workbook()
    #             self.error_sh=self.error_wb.active

    #             print('2')
    #             self.list_to_pool=[]
    #             for i, wb in enumerate(self.source.glob('*.xls*'), start=1):
    #                 self.temp_dict={'wb':wb, 'i':i}
    #                 self.list_to_pool.append(self.temp_dict)
                
    #             self.max_core=psutil.cpu_count(logical=False)
    #             print('Total physical cores: ', self.max_core)
    #             p=Pool(self.max_core-1)
    #             print('after pool declaration')
    #             self.result_of_pool=p.map(self.pooler, self.list_to_pool)
    #             print('after pool map')
    #             p.close()
    #             p.join()

    #         #  Working without Pool
    #             # for i, wb in enumerate(self.source.glob('*.xls*'), start=1):
    #             #     print('\n' + 'Started working on: '.ljust(25, ' '), wb.name)
    #             #     # First Check user's replace or not selection, handle the request upon.             
    #             #     self.filename=str(wb.stem)+ '.pdf'
    #             #     self.dest_full_name= r'' + str(self.destination / self.filename) + ''
    #             #     # If our suggested destination path does not exist, then create
    #             #     if not self.destination.exists():
    #             #         self.destination.mkdir(parents=True, exist_ok=True)
    #             #     # print('2.1')
    #             #     # If user does not want to replace
    #             #     # Keep old files and skip current facture
    #             #     if self.rbtn_value.get()==1: # Wants to keep old, does not want to replace
    #             #         if Path(self.dest_full_name).exists():
    #             #             print('Already exists, skipping: '.ljust(25, ' '), self.filename)
    #             #             continue
    #             #     # print('2.2 after replace check')
    #             #     # print('3')
    #             #     self.wb_analyzer=AnalyzeWB(wb=wb, error_sh=self.error_sh, error_row=i, xl=self.xl)
    #             #     # print('3.1')
    #             #     self.error_sh.cell(row=i, column=1, value=str(wb.stem))
    #             #     # print('4')
    #             #     if  self.wb_analyzer.is_proper_workbook() and self.wb_analyzer.is_facture_wb():
    #             #         # print('6')

    #             #         self.wb_current=self.wb_analyzer.get_wb()
    #             #         self.shfacture=self.wb_analyzer.get_facture_sheet()
    #             #         self.shannexe=self.wb_analyzer.get_annexe_sheet()

    #             #         # print('8')
    #             #         # Select both sheets (Facture and Annexe)
    #             #         self.wb_current.Worksheets([self.shfacture.Name, self.shannexe.Name]).Select()

    #             #         # Analyze Facture Sheet
    #             #         self.shfacture_analyze=AnalyzeFactureSheet(sh=self.shfacture, error_sh=self.error_sh, error_row=i)
    #             #         # Analyze Annexe Sheet
    #             #         self.shannexe_analyze=AnalyzeAnnexeSheet(sh=self.shannexe, error_sh=self.error_sh, error_row=i)

    #             #         # Export to PDF
    #             #         try:
    #             #             # print('destination address', self.dest_full_name)
    #             #             self.xl.ActiveSheet.ExportAsFixedFormat(0, self.dest_full_name)
    #             #             self.error_sh.cell(row=i, column=2, value=str(wb.stem) + ' converted successfully :)')
    #             #             print('Converted successfully: '.ljust(25, ' ') + str(wb.stem))
    #             #         except:
    #             #             print('Couldn\'t convert to PDF. It looks like PDF with same name is open, \n Please close and try again.')
    #             #             self.error_sh.cell(row=i, column=2, value=str(wb.stem) + ' Error when converting :(')

    #             #         # Close workbooks as soon pdf is created.
    #             #         self.wb_current.Saved=True
    #             #         self.wb_current.Close()

    #         except Exception as e:
    #             print('Exception here: ', e)

    #             # xl.Calculation = xlCalculationAutomatic

    #             try:
    #                 print('Exception again: ', e)

    #                 self.wb_current.Saved=True
    #                 self.wb_current.Close()
    #             except:
    #                 pass 

    #             # self.xl.Quit()

    #         finally:
    #             # self.xl.ScreenUpdating = True
    #             # self.xl.EnableEvents = True
    #             # self.xl.DisplayAlerts = True
    #             # self.xl.AskToUpdateLinks = True
    #             # xl.Calculation = xlCalculationAutomatic

    #             # self.xl.Quit()
    #             pass
            
    #         self.error_wb.save(Path.cwd() / 'Here are the errors.xlsx')
    #         self.finish=time.perf_counter()
    #         print(f'\nTotal time spent for PDF convertion: {round(self.finish-self.start, 2)}')

    #     def pooler(self, wb_info):

    #         try:
    #             self.xl=win32.gencache.EnsureDispatch('Excel.Application')
                    
    #             self.xl.ScreenUpdating = False
    #             self.xl.EnableEvents = False
    #             self.xl.DisplayAlerts = False
    #             self.xl.AskToUpdateLinks = False
                

    #             self.xl.Visible=False


    #             #   for i, wb in enumerate(self.source.glob('*.xls*'), start=1):
    #             # self.wb=args[0]['wb']
    #             # self.i=args[0]['i']
    #             self.wb=wb_info['wb']
    #             self.i=wb_info['i']
    #             # self.wb_current=''
    #             print('\n' + 'Started working on: '.ljust(25, ' '), self.wb.name)
    #             # First Check user's replace or not selection, handle the request upon.             
    #             self.filename=str(self.wb.stem)+ '.pdf'
    #             self.dest_full_name= r'' + str(self.destination / self.filename) + ''
    #             # If our suggested destination path does not exist, then create
    #             if not self.destination.exists():
    #                 self.destination.mkdir(parents=True, exist_ok=True)
    #             # print('2.1')
    #             # If user does not want to replace
    #             # Keep old files and skip current facture
    #             if self.selection==1: # Wants to keep old, does not want to replace
    #                 if Path(self.dest_full_name).exists():
    #                     print('Already exists, skipping: '.ljust(25, ' '), self.filename)
    #                     return {}
    #             # print('2.2 after replace check')
    #             # print('3')
    #             self.wb_analyzer=AnalyzeWB(wb=self.wb, error_sh=self.error_sh, error_row=self.i, xl=self.xl)
    #             # print('3.1')
    #             self.error_sh.cell(row=self.i, column=1, value=str(self.wb.stem))
    #             # print('4')
    #             if  self.wb_analyzer.is_proper_workbook() and self.wb_analyzer.is_facture_wb():
    #                 # print('6')

    #                 self.wb_current=self.wb_analyzer.get_wb()
    #                 self.shfacture=self.wb_analyzer.get_facture_sheet()
    #                 self.shannexe=self.wb_analyzer.get_annexe_sheet()

    #                 # print('8')
    #                 # Select both sheets (Facture and Annexe)
    #                 self.wb_current.Worksheets([self.shfacture.Name, self.shannexe.Name]).Select()

    #                 # Analyze Facture Sheet
    #                 self.shfacture_analyze=AnalyzeFactureSheet(sh=self.shfacture, error_sh=self.error_sh, error_row=self.i)
    #                 # Analyze Annexe Sheet
    #                 self.shannexe_analyze=AnalyzeAnnexeSheet(sh=self.shannexe, error_sh=self.error_sh, error_row=self.i)

    #                 # Export to PDF
    #                 try:
    #                     # print('destination address', self.dest_full_name)
    #                     self.xl.ActiveSheet.ExportAsFixedFormat(0, self.dest_full_name)
    #                     self.error_sh.cell(row=self.i, column=2, value=str(self.wb.stem) + ' converted successfully :)')
    #                     print('Converted successfully: '.ljust(25, ' ') + str(self.wb.stem))
    #                 except:
    #                     print('Couldn\'t convert to PDF. It looks like PDF with same name is open, \n Please close and try again.')
    #                     self.error_sh.cell(row=self.i, column=2, value=str(self.wb.stem) + ' Error when converting :(')

                    
    #         except Exception as e:
    #             print('Exception inside pool function: ', e )
    #             try:
    #                 print('Exception inside pool function again: ', e )
    #                 self.wb_current.Saved=True
    #                 self.wb_current.Close()
    #             except:
    #                 pass 
                
    #             self.xl.ScreenUpdating = True
    #             self.xl.EnableEvents = True
    #             self.xl.DisplayAlerts = True
    #             self.xl.AskToUpdateLinks = True

    #             self.xl.Quit()
    #         finally:
    #             # Close workbooks as soon pdf is created.
    #             self.wb_current.Saved=True
    #             self.wb_current.Close()

    #             self.xl.ScreenUpdating = True
    #             self.xl.EnableEvents = True
    #             self.xl.DisplayAlerts = True
    #             self.xl.AskToUpdateLinks = True

    #             self.xl.Quit()

            
    #         return {}

