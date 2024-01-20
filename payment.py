import pandas as pd
import time
import os
import gc
import sys

import numpy as np
from pathlib import Path
import shutil
from PyPDF2 import PdfFileReader, PdfFileWriter, PdfFileMerger
import psutil
from pdfrw import PdfReader, PdfWriter, IndirectPdfDict
from fpdf import FPDF
import openpyxl
from openpyxl import Workbook

from last_modife import LastModifeFinder
from helper import ExcelAnalyzer, next_available_name,next_available_folder_name, available_ram, routage_splitter, tds_splitter, decl_splitter, facture_miner


class Payment:
    def __init__(self, *args, **kwargs):
        # Situation YGT Excel File
        self.situation_file = kwargs['situation_file']
        # MT Folder
        self.source_parent = kwargs['source_parent']
        # General Invoice Excel File
        self.general_invoice_file = kwargs['general_invoice_file']
        # Destination Folder will be next to Situation File

        self.destination =next_available_folder_name (Path(self.situation_file).parent / 'Payment Documentations')
        self.destination.mkdir()

        # Dictionaries needed throughout runtime
        self.dict_to_copy = {}
        self.dict_situation = {}
        self.dict_decl_splitted = {}
        self.dict_tds_splitted = {}
        self.not_found_docs = {
            'boq': [],
            'facture': [],
            'routage': [],
            'decl': [],
            'tds': [],
            'coo': [],
            'destination': [],
            'other': []
        }
        # User's wishes
        self.want_facture = True
        self.want_routage = True
        self.want_decl = True
        self.want_tds = True
        self.want_coo = True

        self.want_whole_decl = False
        self.want_decl_page = False
        self.want_whole_tds = False
        self.want_tds_page = False

        # Prevent duplicate search with these sets
        self.set_searched_factures = set()

        self.assign_source_files_to_dataframe()

    def assign_source_files_to_dataframe(self):
        facture_folder = r'1-FACTURE'
        routage_folder = r'2-CMR'
        decl_folder = r'3-DECLARATION'
        tds_folder = r'4-TDS'
        coo_folder = r'5-CO'

        facture_path = self.source_parent / facture_folder
        routage_path = self.source_parent / routage_folder
        decl_path = self.source_parent / decl_folder
        tds_path = self.source_parent / tds_folder
        coo_path = self.source_parent / coo_folder

        # Create modife finder to use later
        self.last_modife_finder = LastModifeFinder(parent_path=facture_path)

        # Convert paths to list
        list_facture_path = [str(x) for x in facture_path.rglob('*.pdf')]
        list_routage_path = [str(x) for x in routage_path.rglob('*.pdf')]
        list_decl_path = [str(x) for x in decl_path.rglob('*.pdf')]
        list_tds_path = [str(x) for x in tds_path.rglob('*.pdf')]
        list_coo_path = [str(x) for x in coo_path.rglob('*.pdf')]

        # Convert lists to Dataframe.
        self.df_facture = pd.DataFrame(list_facture_path, columns=['path'])
        self.df_routage = pd.DataFrame(list_routage_path, columns=['path'])
        self.df_decl = pd.DataFrame(list_decl_path, columns=['path'])
        self.df_tds = pd.DataFrame(list_tds_path, columns=['path'])
        self.df_coo = pd.DataFrame(list_coo_path, columns=['path'])

        # Set to None
        list_facture_path = None
        list_routage_path = None
        list_decl_path = None
        list_tds_path = None
        list_coo_path = None

    def assign_situation_file_to_dictionary(self):
        excel_analyzer = ExcelAnalyzer(excel=self.situation_file)
        excel_analyzer.analyze_for_ygt()
        self.x_col = 0
        self.lot_col = 2
        self.boq_code_col = 3
        self.facture_col = 5
        self.decl_no_col = 13
        self.decl_haryt_col = 14
        if self.want_tds_page:
            self.tds_no_col = 42
            self.tds_page_col = 43

        if excel_analyzer.what_is_it == 'ygt':
            df_situation = excel_analyzer.df_situation
            rows_count = df_situation.shape[0]
            print(rows_count)
            print('Assigning situation excel to dictionary...')
            for i in range(5, rows_count):
                if str(df_situation.iloc[i, self.x_col]).lower() == 'x' and \
                    not str(df_situation.iloc[i, self.lot_col]).lower() == 'lot' and \
                        not str(df_situation.iloc[i, self.facture_col]).lower() == '':
                    boq_code = str(str(df_situation.iloc[i, self.boq_code_col]).strip()).replace('\n', ' ').replace('\"', '-').replace(
                        '/', '-').replace(':', '-').replace('*', '-').replace('?', '-').replace('>', '-').replace('<', '-').replace('|', '-')
                    facture_no = str(str(df_situation.iloc[i, self.facture_col]).strip()).replace('\n', ' ').replace('\"', '-').replace(
                        '/', '-').replace(':', '-').replace('*', '-').replace('?', '-').replace('>', '-').replace('<', '-').replace('|', '-')
                    if str(facture_no) == 'nan':
                        self.not_found_docs['facture'].append(
                            'NAN facture_no in row %s' % (i + 2))
                        continue

                    decl_no = str(str(df_situation.iloc[i, self.decl_no_col]).strip()).replace('\n', ' ').replace('\"', '-').replace(
                        '/', '-').replace(':', '-').replace('*', '-').replace('?', '-').replace('>', '-').replace('<', '-').replace('|', '-')
                    # print(i)
                    # print(boq_code)
                    # print(decl_no)
                    # decl_no = str(str(str(decl_no).replace('/', '-')).strip()).replace('\n', ' ')
                    decl_haryt_no = str(str(df_situation.iloc[i, self.decl_haryt_col]).strip()).replace('\n', ' ').replace(
                        '\"', '-').replace('/', '-').replace(':', '-').replace('*', '-').replace('?', '-').replace('>', '-').replace('<', '-').replace('|', '-')
                    # if decl_haryt_no=='CMR':
                    #         print('stop')
                    decl_page = ''
                    if str(decl_haryt_no).isdigit():
                        # if decl_haryt_no=='CMR':
                        #     print('stop')
                        if int(decl_haryt_no) > 0:
                            # print('haryt no ', decl_haryt_no)
                            decl_haryt_no = int(decl_haryt_no)
                            decl_page = self.find_decl_page(
                                haryt_no=decl_haryt_no)
                            # print('page no ', decl_page)
                    else:
                        # test
                        # print(f'Haryt no {decl_haryt_no} is not a digit')
                        self.not_found_docs['decl'].append(
                            'Haryt no is not a number in row %s' % (i + 2))
                        pass

                    if self.want_tds_page:
                        tds_no = str(
                            df_situation.iloc[i, self.tds_no_col]).strip()
                        tds_page = str(
                            df_situation.iloc[i, self.tds_page_col]).strip()

                    # Handle empty BOQ codes
                    if str(boq_code) == 'nan':
                        self.not_found_docs['boq'].append(
                            'NAN boq in row %s' % (i + 2))
                        continue
                    if len(str(boq_code)) < 1 or str(boq_code) == '0':
                        self.not_found_docs['boq'].append(
                            'Empty boq in row %s' % (i + 2))
                        continue

                    # Handle empty Facture numbers
                    if len(str(facture_no)) < 1 or str(facture_no) == '0':
                        self.not_found_docs['facture'].append(
                            'Empty facture_no in row %s' % (i + 2))
                        continue

                    # Add boq_code key if there is no any
                    try:
                        self.dict_situation[boq_code]
                    except:
                        self.dict_situation[boq_code] = {}
                    # Add facture_no sub key if there is no
                    try:
                        self.dict_situation[boq_code][facture_no]
                    except:
                        self.dict_situation[boq_code][facture_no] = {}
                        # print('facture no is dictionary after exception')

                    # Add declaration_no, if there is declaration_no already then add new page
                    # Check if declaration and declaration haryt number is empty or not
                    if len(str(decl_no)) > 0:
                        if self.want_whole_decl or self.want_decl_page and len(str(decl_page)) > 0 and not str(decl_page) == '0':
                            try:
                                self.dict_situation[boq_code][facture_no][decl_no]
                                self.dict_situation[boq_code][facture_no][decl_no].add(
                                    decl_page)
                            except:
                                self.dict_situation[boq_code][facture_no][decl_no] = set(
                                )
                                self.dict_situation[boq_code][facture_no][decl_no].add(
                                    decl_page)
                        else:
                            self.not_found_docs['decl'].append(
                                'Empty declaration haryt no in row %s' % (i + 2))
                    else:
                        self.not_found_docs['decl'].append(
                            'Empty declaration no in row %s' % (i + 2))

                    # Add TDS no and page if user wishes and if they are not empty, if empty then add it to not_found_dict
                    if self.want_tds_page:
                        if len(str(tds_page)) > 0 and len(str(tds_no)) > 0 and not str(tds_page) == '0' and not str(tds_no) == '0':
                            try:
                                self.dict_situation[boq_code][facture_no][tds_no]
                                self.dict_situation[boq_code][facture_no][tds_no].add(
                                    tds_page)
                            except:
                                self.dict_situation[boq_code][facture_no][tds_no] = set(
                                )
                                self.dict_situation[boq_code][facture_no][tds_no].add(
                                    tds_page)
                        else:
                            self.not_found_docs['tds'].append(
                                'Empty TDS or TDS page number in row %s' % (i + 2))

            df_situation = None
        else:  # If not YGT
            print('Situation file is not YGT, please select proper situation file.')
            pass

    def assign_general_invoice_to_dataframe(self):
        self.df_general_invoice = pd.read_excel(
            self.general_invoice_file, sheet_name='GENERAL INVOICE')

    def start_searching(self):
        for boq_code in list(self.dict_situation):
            self.boq_code = boq_code
            for facture_no in list(self.dict_situation[boq_code]):
                self.is_facture_no_from_trash = False
                self.possible_facture_no_from_trash = ''
                self.facture_no_contains_trojan = False
                print('Searching documents of {}'.format(facture_no))
                if 'nan' in str(facture_no):
                    continue

                # FACTURE
                if self.want_facture:
                    self.facture_no = self.dict_situation[boq_code][facture_no]
                    self.find_facture(boq_code=self.boq_code,
                                      facture_no=facture_no)
                    # print('Inside start_searching: ', facture_no)

                # Sometimes factures contains brackets that has no closing end, that causes error.
                if self.facture_no_contains_trojan:
                    continue

                if self.is_facture_no_from_trash:
                    self.dict_situation[boq_code][self.possible_facture_no_from_trash[0]
                                                  ] = self.dict_situation[boq_code].pop(facture_no)
                    facture_no = self.possible_facture_no_from_trash[0]

                # Routage
                if self.want_routage:
                    # print('will find routage_no according to facture_no: ', facture_no)
                    try:
                        # Since facture number is written strange sometimes, it raises error. If it is the case then skip this facture
                        self.routage = self.df_general_invoice[self.df_general_invoice.iloc[:, 0].str.match(
                            facture_no, case=False, na=False)]
                    except:
                        print(
                            'Error with facture, when trying to find routage - {}'.format(facture_no))
                        continue

                    self.routage = self.routage.iloc[:, 1].tolist()
                    for routage_no in self.routage:
                        # print('will call find_routage for: ', routage_no)
                        self.find_routage(
                            boq_code=boq_code, facture_no=facture_no, routage_no=routage_no)

                # TDS
                if self.want_tds:
                    if self.want_whole_tds:
                        # print('will find tds_no according to facture_no: ', facture_no)
                        self.tds = self.df_general_invoice[self.df_general_invoice.iloc[:, 0].str.match(
                            facture_no, case=False, na=False)]
                        self.tds = self.tds.iloc[:, 3].tolist()
                        for tds_no in self.tds:
                            # print('will call find_tds for: ', tds_no)
                            self.find_tds(boq_code=boq_code,
                                          facture_no=facture_no, tds_no=tds_no)
                    elif self.want_tds_page:
                        for tds_no in self.dict_situation[boq_code][facture_no]:
                            # Each decl_no of dict_situation keeps necessary pages as list
                            pages = self.dict_situation[boq_code][facture_no][tds_no]
                            # print('splitting: ' , decl_no)
                            self.find_tds_page_by_page(
                                boq_code=boq_code, facture_no=facture_no, tds_no=tds_no, tds_pages=pages)

                # COO
                if self.want_coo:
                    self.coo_no = facture_no
                    # print('will call find_tds for: ', tds_no)
                    self.find_coo(boq_code=boq_code,
                                  facture_no=facture_no, coo_no=self.coo_no)

                # DECLARATION
                for decl_no in self.dict_situation[boq_code][facture_no]:
                    # Each decl_no of dict_situation keeps necessary pages as list
                    pages = self.dict_situation[boq_code][facture_no][decl_no]
                    # print('splitting: ' , decl_no)
                    self.find_decl(
                        boq_code=boq_code, facture_no=facture_no, decl_no=decl_no, decl_pages=pages)

                # Update searched factures set
                self.set_searched_factures.add(facture_no)

    def find_facture(self, boq_code, facture_no):
        if not facture_no in self.set_searched_factures:
            # print('Searching: ', facture_no)
            pattern1 = r'.*?(\D|\b)' + str(facture_no) + r'\D.*[pP][dD][fF]$'
            pattern2 = r'.*?[^-]' + str(facture_no) + r'\D.*[pP][dD][fF]$'
            try:
                result = self.df_facture[self.df_facture['path'].str.match(
                    pattern1, case=False, na=False)]
                result = result[result['path'].str.match(
                    pattern2, case=False, na=False)]
            except:
                print('Error happened when searching facture (probably there is space at end of facture in cell) {}'.format(
                    facture_no))
                # NOT FOUND
                self.not_found_docs['facture'].append(facture_no)
                self.facture_no_contains_trojan = True
                return

            # Convert found facture paths to facture names list
            file_name_list = []
            for file_name in result['path'].tolist():
                file_name_list.append(file_name)

            # If file_name_list is empty, then it is possible that facture name is written trash format
            # If it is the case, we pull pure facture number from trash format and append it to the list
            if len(file_name_list) == 0:
                self.possible_facture_no_from_trash = facture_miner(facture_no)
                if len(self.possible_facture_no_from_trash) == 1:
                    print(f'Trash {facture_no}')
                    facture_no = self.possible_facture_no_from_trash[0]
                    print(f'Mined from trash {facture_no}')
                    self.is_facture_no_from_trash = True
                    pattern11 = r'.*?(\D|\b)' + str(
                        self.possible_facture_no_from_trash[0]) + r'\D.*[pP][dD][fF]$'
                    pattern21 = r'.*?[^-]' + str(
                        self.possible_facture_no_from_trash[0]) + r'\D.*[pP][dD][fF]$'
                    result = self.df_facture[self.df_facture['path'].str.match(
                        pattern11, case=False, na=False)]
                    result = result[result['path'].str.match(
                        pattern21, case=False, na=False)]
                    file_name_list = []
                    for file_name in result['path'].tolist():
                        file_name_list.append(file_name)
                    # file_name_list.append(self.possible_facture_no_from_trash[0])

            # Find last modife, if search result is more than one file
            if len(file_name_list) > 1:
                self.last_modife_finder.find_last_modife(
                    facture_no=facture_no, file_name_list=file_name_list)
                self.last_modife = self.last_modife_finder.last_modife
                if self.last_modife == 'none':
                    temp_len = []
                    for temp_file in file_name_list:
                        temp_len.append(len(temp_file))

                    self.last_modife = file_name_list[np.argmax(temp_len)]

            elif len(file_name_list) == 1:
                self.last_modife = file_name_list[0]

            # If file_name_list is empty then there is no match
            else:
                # NOT FOUND
                self.not_found_docs['facture'].append(facture_no)
                self.last_modife = ''
                # Add facture_no to to_copy dictionary, even if it is not found
                try:
                    self.dict_to_copy[facture_no]
                except:
                    self.dict_to_copy[facture_no] = {}

            if not self.last_modife == '':
                # Add facture_no to to_copy dictionary
                try:
                    self.dict_to_copy[facture_no]
                except:
                    self.dict_to_copy[facture_no] = {}

                # Add facture_pdf name to to_copy dictionary
                try:
                    self.dict_to_copy[facture_no]['facture_pdf']
                    self.dict_to_copy[facture_no]['facture_pdf'].append(
                        self.last_modife)
                except:
                    self.dict_to_copy[facture_no]['facture_pdf'] = list()
                    self.dict_to_copy[facture_no]['facture_pdf'].append(
                        self.last_modife)

                # print('adding destination to: ', list(self.destination / boq_code / facture_no))
                # Update destination address for found facture
                try:
                    self.dict_to_copy[facture_no]['destination']
                    self.dict_to_copy[facture_no]['destination'].add(
                        self.destination / boq_code / facture_no)
                except:
                    self.dict_to_copy[facture_no]['destination'] = set()
                    self.dict_to_copy[facture_no]['destination'].add(
                        self.destination / boq_code / facture_no)
        else:  # Update destination address only
            # print('Already searched: ', facture_no)
            try:
                self.dict_to_copy[facture_no]['destination']
                self.dict_to_copy[facture_no]['destination'].add(
                    self.destination / boq_code / facture_no)
            except:
                self.dict_to_copy[facture_no]['destination'] = set()
                self.dict_to_copy[facture_no]['destination'].add(
                    self.destination / boq_code / facture_no)

    def find_routage(self, boq_code, facture_no, routage_no):
        list_found_routage = list()
        if not facture_no in self.set_searched_factures:
            # print('Searching: ', routage_no)
            if '/' in routage_no:
                routage_no = str(routage_no).replace('/','-')
            pattern1 = r'.*?' + str(routage_no) + r'\D.*[pP][dD][fF]$'
            found_routage = self.df_routage[self.df_routage['path'].str.match(
                pattern1, case=False, na=False)]

            # print('found_routage: ', found_routage)
            # print('self.df_routage[path]:, ', self.df_routage['path'])
            # there might be several routage found, but only 1 is enough

            if len(found_routage['path'].tolist()) == 0:
                self.partial_routage_list = routage_splitter(routage_no)
                for partial_routage in self.partial_routage_list:
                    pattern_partial = r'.*?' + \
                        str(partial_routage) + r'\D.*[pP][dD][fF]$'
                    found_routage = self.df_routage[self.df_routage['path'].str.match(
                        pattern_partial, case=False, na=False)]
                    if len(found_routage['path'].tolist()) == 0:
                        # NOT FOUND
                        self.not_found_docs['routage'].append(routage_no)
                        # Add facture_no to to_copy dictionary even if it is not found
                        try:
                            self.dict_to_copy[facture_no]
                        except:
                            self.dict_to_copy[facture_no] = {}
                    else:
                        list_found_routage.append(
                            found_routage['path'].tolist()[0])

                # # Add function that will detect routage numbers from trash and find them seperately
                # return
            else:
                list_found_routage.append(found_routage['path'].tolist()[0])

            for rtg in list_found_routage:
                # Add facture_no to to_copy dictionary
                try:
                    self.dict_to_copy[facture_no]
                except:
                    self.dict_to_copy[facture_no] = {}

                # Add routage_pdf name to to_copy dictionary
                try:
                    self.dict_to_copy[facture_no]['routage_pdf']
                    self.dict_to_copy[facture_no]['routage_pdf'].append(rtg)
                except:
                    self.dict_to_copy[facture_no]['routage_pdf'] = list()
                    self.dict_to_copy[facture_no]['routage_pdf'].append(rtg)

                # print('adding destination to: ', list(self.destination / boq_code / facture_no))
                # Update destination address for found facture
                try:
                    self.dict_to_copy[facture_no]['destination']
                    self.dict_to_copy[facture_no]['destination'].add(
                        self.destination / boq_code / facture_no)
                except:
                    self.dict_to_copy[facture_no]['destination'] = set()
                    self.dict_to_copy[facture_no]['destination'].add(
                        self.destination / boq_code / facture_no)
        else:  # Update destination address only
            # print('Already searched: ', facture_no)
            try:
                self.dict_to_copy[facture_no]['destination']
                self.dict_to_copy[facture_no]['destination'].add(
                    self.destination / boq_code / facture_no)
            except:
                self.dict_to_copy[facture_no]['destination'] = set()
                self.dict_to_copy[facture_no]['destination'].add(
                    self.destination / boq_code / facture_no)

    def find_tds(self, boq_code, facture_no, tds_no):
        list_found_tds = list()
        if not facture_no in self.set_searched_factures:
            # print('Searching: ', tds_no)
            pattern = r'.*?' + str(tds_no) + r'\D.*[pP][dD][fF]$'
            found_tds = self.df_tds[self.df_tds['path'].str.match(
                pattern, case=False, na=False)]

            # print('found_tds: ', found_tds)
            # print('self.df_tds[path]:, ', self.df_tds['path'])

            # There might be several tds found, but only 1 is enough
            if len(found_tds['path'].tolist()) == 0:
                self.partial_tds_list = tds_splitter(tds_no)
                for partial_tds in self.partial_tds_list:
                    pattern_partial = r'.*?' + \
                        str(partial_tds) + r'\D.*[pP][dD][fF]$'
                    found_tds = self.df_tds[self.df_tds['path'].str.match(
                        pattern_partial, case=False, na=False)]
                    if len(found_tds['path'].tolist()) == 0:
                        # NOT FOUND
                        self.not_found_docs['tds'].append(tds_no)
                        # Add facture_no to to_copy dictionary even if it is not found
                        try:
                            self.dict_to_copy[facture_no]
                        except:
                            self.dict_to_copy[facture_no] = {}
                    else:
                        list_found_tds.append(found_tds['path'].tolist()[0])
            else:
                list_found_tds.append(found_tds['path'].tolist()[0])

            # Loop through found tds list, and assign them into dictionary with destination
            for tds in list_found_tds:
                # Add facture_no to to_copy dictionary
                try:
                    self.dict_to_copy[facture_no]
                except:
                    self.dict_to_copy[facture_no] = {}

                # Add tds_pdf name to to_copy dictionary
                try:
                    self.dict_to_copy[facture_no]['tds_pdf']
                    self.dict_to_copy[facture_no]['tds_pdf'].append(tds)
                except:
                    self.dict_to_copy[facture_no]['tds_pdf'] = list()
                    self.dict_to_copy[facture_no]['tds_pdf'].append(tds)

                # print('adding destination to: ', list(self.destination / boq_code / facture_no))
                # Update destination address for found facture
                try:
                    self.dict_to_copy[facture_no]['destination']
                    self.dict_to_copy[facture_no]['destination'].add(
                        self.destination / boq_code / facture_no)
                except:
                    self.dict_to_copy[facture_no]['destination'] = set()
                    self.dict_to_copy[facture_no]['destination'].add(
                        self.destination / boq_code / facture_no)
        else:  # Update destination address only
            # print('Already searched: ', facture_no)
            try:
                self.dict_to_copy[facture_no]['destination']
                self.dict_to_copy[facture_no]['destination'].add(
                    self.destination / boq_code / facture_no)
            except:
                self.dict_to_copy[facture_no]['destination'] = set()
                self.dict_to_copy[facture_no]['destination'].add(
                    self.destination / boq_code / facture_no)

    def find_tds_page_by_page(self, boq_code, facture_no, tds_no, tds_pages):
        list_found_tds = list()
        pattern = r'.*?' + str(tds_no) + r'\D.*[pP][dD][fF]$'
        found_tds = self.df_tds[self.df_tds['path'].str.match(
            pattern, case=False, na=False)]

        # There might be several tds found, but only 1 is enough
        if len(found_tds['path'].tolist()) == 0:
            self.partial_tds_list = tds_splitter(tds_no)
            for partial_tds in self.partial_tds_list:
                pattern_partial = r'.*?' + \
                    str(partial_tds) + r'\D.*[pP][dD][fF]$'
                found_tds = self.df_tds[self.df_tds['path'].str.match(
                    pattern_partial, case=False, na=False)]
                if len(found_tds['path'].tolist()) == 0:
                    # NOT FOUND
                    self.not_found_docs['tds'].append(tds_no)
                    pass
                else:
                    list_found_tds.append(found_tds['path'].tolist()[0])
        else:
            list_found_tds.append(found_tds['path'].tolist()[0])

        # Loop through found tds list, split them page by page, and assign them into dictionary with destination
        for tds_path in list_found_tds:
            # print('inside split TDS')
            try:
                # If it is already splitted
                self.dict_tds_splitted[tds_path]
                # Add Destination
                for situation_page in tds_pages:
                    try:
                        self.dict_tds_splitted[tds_path][situation_page]['destination']
                        self.dict_tds_splitted[tds_path][situation_page]['destination'].add(
                            self.destination / boq_code / facture_no)
                    except:
                        self.dict_tds_splitted[tds_path][situation_page]['destination'] = set(
                        )
                        self.dict_tds_splitted[tds_path][situation_page]['destination'].add(
                            self.destination / boq_code / facture_no)
            except:
                # If not splitted then split
                reader = PdfReader(tds_path)
                writer_cover = PdfWriter()
                writer_cover.addPage(reader.pages[0])
                # total_pages=decl.numPages
                total_pages = len(reader.pages)
                # print('total pages: ', total_pages)
                for page in range(total_pages):
                    writer = PdfWriter()
                    writer.addPage(reader.pages[0])
                    writer.addPage(reader.pages[page])

                    # Add TDS into splitted dictionary
                    try:
                        self.dict_tds_splitted[tds_path]
                    except:
                        self.dict_tds_splitted[tds_path] = {}

                    # Add pdf page into dictionary
                    try:
                        self.dict_tds_splitted[tds_path][page+1]
                    except:
                        self.dict_tds_splitted[tds_path][page+1] = {}
                        self.dict_tds_splitted[tds_path][page +
                                                         1]['pdf_cover'] = writer_cover
                        self.dict_tds_splitted[tds_path][page +
                                                         1]['pdf'] = writer

                    # Add Destination
                    for situation_page in tds_pages:
                        if not page+1 == situation_page:
                            continue
                        try:
                            self.dict_tds_splitted[tds_path][situation_page]['destination']
                            self.dict_tds_splitted[tds_path][situation_page]['destination'].add(
                                self.destination / boq_code / facture_no)
                        except:
                            self.dict_tds_splitted[tds_path][situation_page]['destination'] = set(
                            )
                            self.dict_tds_splitted[tds_path][situation_page]['destination'].add(
                                self.destination / boq_code / facture_no)

                reader = None
                writer = None
                writer_cover = None

    def find_coo(self, boq_code, facture_no, coo_no):
        if not facture_no in self.set_searched_factures:
            # print('Searching: ', facture_no)
            pattern1 = r'.*?(\D|\b)' + str(facture_no) + r'\D.*[pP][dD][fF]$'
            pattern2 = r'.*?[^-]' + str(facture_no) + r'\D.*[pP][dD][fF]$'
            try:
                result = self.df_coo[self.df_coo['path'].str.match(
                    pattern1, case=False, na=False)]
                result = result[result['path'].str.match(
                    pattern2, case=False, na=False)]
            except:
                print('Error when searching {}', coo_no)
                # NOT FOUND
                self.not_found_docs['coo'].append(coo_no)
                return

            # there might be several COO found, but only 1 is enough
            if len(result['path'].tolist()) == 0:
                # NOT FOUND
                self.not_found_docs['coo'].append(coo_no)
                # Add not found facture as a key of dict, even if it is not found
                try:
                    self.dict_to_copy[facture_no]
                except:
                    self.dict_to_copy[facture_no] = {}

            # result=result['path'].tolist()[0]
            # Here take all COOs which are found.

            for rslt in result['path'].tolist():
                try:
                    self.dict_to_copy[facture_no]
                except:
                    self.dict_to_copy[facture_no] = {}
                # Add facture_pdf name to to_copy dictionary
                try:
                    self.dict_to_copy[facture_no]['coo_pdf']
                    self.dict_to_copy[facture_no]['coo_pdf'].append(rslt)
                except:
                    self.dict_to_copy[facture_no]['coo_pdf'] = list()
                    self.dict_to_copy[facture_no]['coo_pdf'].append(rslt)

                # print('adding destination to: ', list(self.destination / boq_code / facture_no))
                # Update destination address for found facture
                try:
                    self.dict_to_copy[facture_no]['destination']
                    self.dict_to_copy[facture_no]['destination'].add(
                        self.destination / boq_code / facture_no)
                except:
                    self.dict_to_copy[facture_no]['destination'] = set()
                    self.dict_to_copy[facture_no]['destination'].add(
                        self.destination / boq_code / facture_no)
        else:  # Update destination address only
            # print('Already searched: ', facture_no)
            try:
                self.dict_to_copy[facture_no]['destination']
                self.dict_to_copy[facture_no]['destination'].add(
                    self.destination / boq_code / facture_no)
            except:
                self.dict_to_copy[facture_no]['destination'] = set()
                self.dict_to_copy[facture_no]['destination'].add(
                    self.destination / boq_code / facture_no)

    def find_decl(self, boq_code, facture_no, decl_no, decl_pages):
        # list_found_decl = list()
        # print('Searching: ', decl_no)
        pattern = r'.*?' + str(decl_no) + r'\D.*[pP][dD][fF]$'
        found_decl = self.df_decl[self.df_decl['path'].str.match(
            pattern, case=False, na=False)]
        # There might be several declaration found, but only 1 is enough
        if len(found_decl['path'].tolist()) == 0:
            # NOT FOUND
            self.not_found_docs['decl'].append(decl_no)
            return
        # There will be one declaration number per material
            # So that, you don't have to search for multiple declaration
            # In case, multiple declaration be matter, then you can use this code below

            #     self.partial_decl_list = decl_splitter(decl_no)
            #     for partial_decl in self.partial_decl_list:
            #         pattern_partial=r'.*?' + str(partial_decl) + r'\D.*[pP][dD][fF]$'
            #         found_decl = self.df_decl[self.df_decl['path'].str.match(pattern_partial, case=False, na=False)]
            #         if len(found_decl['path'].tolist())==0:
            #             pass
            #         else:
            #             list_found_decl.append(found_decl['path'].tolist()[0])
            # else:
            #     list_found_decl.append(found_decl['path'].tolist()[0])

            # Here add function that detects declaration numbers from trash and find them seperately
            #     return

        found_decl = found_decl['path'].tolist()[0]
        # Add facture_no to to_copy dictionary
        # for dcl in list_found_decl:
        try:
            self.dict_to_copy[facture_no]
        except:
            self.dict_to_copy[facture_no] = {}
        # Add decl_pdf name to to_copy dictionary
        try:
            self.dict_to_copy[facture_no]['decl_pdf']
            self.dict_to_copy[facture_no]['decl_pdf'].append(found_decl)
        except:
            self.dict_to_copy[facture_no]['decl_pdf'] = list()
            self.dict_to_copy[facture_no]['decl_pdf'].append(found_decl)
        # Update destination address for found facture
        try:
            self.dict_to_copy[facture_no]['destination']
            self.dict_to_copy[facture_no]['destination'].add(
                self.destination / boq_code / facture_no)
        except:
            self.dict_to_copy[facture_no]['destination'] = set()
            self.dict_to_copy[facture_no]['destination'].add(
                self.destination / boq_code / facture_no)

        # Declaration dictionary to split page by page
        if self.want_decl_page:
            try:
                self.dict_decl_splitted[found_decl]
            except:
                self.dict_decl_splitted[found_decl] = {}
                # Split found declaration
            self.split_declaration(
                path=found_decl, 
                destination=self.destination / boq_code / facture_no, 
                decl_pages=decl_pages)

    def split_declaration(self, path, destination, decl_pages):
        # print('inside split declaraion')
        try:
            # If it is already splitted
            self.dict_decl_splitted[path]
            # Add Destination
            for situation_page in decl_pages:
                try:
                    self.dict_decl_splitted[path][situation_page]['destination']
                    self.dict_decl_splitted[path][situation_page]['destination'].add(
                        destination)
                except:
                    self.dict_decl_splitted[path][situation_page]['destination'] = set(
                    )
                    self.dict_decl_splitted[path][situation_page]['destination'].add(
                        destination)
        except:
            # If not splitted then split
            # temp_pdf_file=open(path, 'rb')
            # decl=PdfFileReader(temp_pdf_file)
            reader = PdfReader(path)
            # total_pages=decl.numPages
            total_pages = len(reader.pages)
            # print('total pages: ', total_pages)
            for page in range(total_pages):
                # seperate_page=PdfFileWriter()
                # decl=PdfFileReader(temp_pdf_file)
                # seperate_page.addPage(decl.getPage(page))
                writer = PdfWriter()
                writer.addPage(reader.pages[page])
                # if it is not first page, then take two page
                if page > 0:
                    try:
                        # decl=PdfFileReader(temp_pdf_file)
                        # seperate_page.addPage(decl.getPage(page+1))
                        writer.addPage(reader.pages[page+1])
                    except:
                        # print('End of declaration page.')
                        pass
                # Add pdf page into dictionary
                try:
                    self.dict_decl_splitted[path][page+1]
                except:
                    self.dict_decl_splitted[path][page+1] = {}
                    # self.dict_decl_splitted[path][page+1]['pdf']=seperate_page
                    self.dict_decl_splitted[path][page+1]['pdf'] = writer
                # Add Destination
                for situation_page in decl_pages:
                    if not page+1 == situation_page:
                        continue
                    try:
                        self.dict_decl_splitted[path][situation_page]['destination']
                        self.dict_decl_splitted[path][situation_page]['destination'].add(
                            destination)
                    except:
                        self.dict_decl_splitted[path][situation_page]['destination'] = set(
                        )
                        self.dict_decl_splitted[path][situation_page]['destination'].add(
                            destination)
            # temp_pdf_file.close()
            # temp_pdf_file = None
            # decl = None
            reader = None
            writer = None

    def find_decl_page(self, haryt_no):
        if haryt_no == 1:
            return 1
        else:
            temp_mod = (haryt_no-1) % 3
            temp_divided = int((haryt_no-1) / 3)
            if temp_mod == 0:
                return temp_divided+1
            else:
                return temp_divided+2

    def copy_files(self):
        gc.collect()
        for facture_no in list(self.dict_to_copy):
            print('Copying documents of: ', facture_no)

            # Destination
            try:
                dst = self.dict_to_copy[facture_no]['destination']
                
            except:
                print('There is no such destionation in facture {}'.format(facture_no))
                self.not_found_docs['destination'].append(
                    'Destination error for {}'.format(facture_no))
                continue

            # Copy Facture
            if self.want_facture:
                try:
                    # print('copying some facture')
                    src_facture = self.dict_to_copy[facture_no]['facture_pdf']
                    for pdf in src_facture:
                        for fldr in dst:
                            if not Path(fldr).exists():
                                Path(fldr).mkdir(parents=True)

                            modified_name = self.add_prename_to_full_file_name(
                                full_path=Path(fldr) / Path(pdf).name, prename='1-')
                            shutil.copy2(pdf, modified_name)
                            # print('copied: ', facture_no)
                except:
                    print('Error when copying facture to destination: ', dst)
                    pass

            # Copy Routage
            if self.want_routage:
                try:
                    # print('copying some routage')
                    src_routage = self.dict_to_copy[facture_no]['routage_pdf']
                    for pdf in src_routage:
                        for fldr in dst:
                            if not Path(fldr).exists():
                                Path(fldr).mkdir(parents=True)

                            modified_name = self.add_prename_to_full_file_name(
                                full_path=Path(fldr) / Path(pdf).name, prename='2-')
                            shutil.copy2(pdf, modified_name)
                            # print('copied: some routage')
                except:
                    print('Error when copying routage to destination: ', dst)
                    pass

            # Copy TDS
            if self.want_whole_tds:
                try:
                    # print('copying some tds')
                    src_tds = self.dict_to_copy[facture_no]['tds_pdf']
                    for pdf in src_tds:
                        for fldr in dst:
                            if not Path(fldr).exists():
                                Path(fldr).mkdir(parents=True)

                            modified_name = self.add_prename_to_full_file_name(
                                full_path=Path(fldr) / Path(pdf).name, prename='4-')
                            shutil.copy2(pdf, modified_name)
                            # print('copied: some tds')
                except:
                    print('Error when copying TDS to destination: ', dst)
                    pass

            if self.want_tds_page:
                for tds in self.dict_tds_splitted:
                    tds_stem = Path(tds).stem
                    tds_suffix = Path(tds).suffix
                    for page in self.dict_tds_splitted[tds]:
                        page_no = ' - {}'.format(page)
                        tds_page_name = '4-{}{}{}'.format(
                            tds_stem, page_no, tds_suffix)
                        # modified_name = self.add_prename_to_full_file_name(full_path=Path(fldr) / Path(pdf).name , prename='2-')
                        # page_pdf_writer=self.dict_tds_splitted[tds][page]['pdf']
                        writer_saver = self.dict_tds_splitted[tds][page]['pdf']
                        # print('in for 2')

                        try:
                            self.dict_tds_splitted[tds][page]['destination']
                            # self.multiple_page_in_one_facture = False
                            for destination in self.dict_tds_splitted[tds][page]['destination']:
                                if not Path(destination).exists():
                                    Path(destination).mkdir(parents=True)
                                tds_page_path = destination / \
                                    str(tds_page_name)
                                # if self.multiple_page_in_one_facture:
                                #     pass
                                with open(tds_page_path, 'wb') as tds_page:
                                    # print('in for 4')
                                    # page_pdf_writer.write(tds_page)
                                    # page_pdf_writer.close()
                                    # page_pdf_writer = None
                                    writer_saver.write(tds_page)
                                    # self.multiple_page_in_one_facture = True
                                    # writer_saver = None
                        except:
                            # print('No need to copy this tds page.')
                            pass

            # Copy COO
            if self.want_coo:
                try:
                    # print('copying some coo')
                    src_coo = self.dict_to_copy[facture_no]['coo_pdf']
                    # print('there is coo file')
                    for pdf in src_coo:
                        for fldr in dst:
                            if not Path(fldr).exists():
                                Path(fldr).mkdir(parents=True)

                            modified_name = self.add_prename_to_full_file_name(
                                full_path=Path(fldr) / Path(pdf).name, prename='5-')
                            shutil.copy2(pdf, modified_name)
                            # print('copied: some coo')
                except:
                    print('Error happened when copyin COO to destination: ', dst)
                    pass

            # Copy DECLARATION
            if self.want_whole_decl:
                try:
                    # print('copying some decl')
                    src_decl = self.dict_to_copy[facture_no]['decl_pdf']
                    for pdf in src_decl:
                        for fldr in dst:
                            if not Path(fldr).exists():
                                Path(fldr).mkdir(parents=True)

                            modified_name = self.add_prename_to_full_file_name(
                                full_path=Path(fldr) / Path(pdf).name, prename='3-')
                            shutil.copy2(pdf, modified_name)
                            # print('copied: some decl')
                except:
                    print(
                        'Error happened when copying declaration to destination: ', dst)
                    pass
            self.dict_to_copy.pop(facture_no)

        # Copy Declaration pages seperately
        if self.want_decl_page:
            for dnumber, decl in enumerate(list(self.dict_decl_splitted)):
                decl_stem = Path(decl).stem
                decl_suffix = Path(decl).suffix
                for page in self.dict_decl_splitted[decl]:
                    page_no = ' - {}'.format(page)
                    decl_page_name = '3-{}{}{}'.format(
                        decl_stem, page_no, decl_suffix)
                    # modified_name = self.add_prename_to_full_file_name(full_path=Path(fldr) / Path(pdf).name , prename='2-')
                    # page_pdf_writer=self.dict_decl_splitted[decl][page]['pdf']
                    writer_saver = self.dict_decl_splitted[decl][page]['pdf']
                    # print('in for 2')

                    try:
                        self.dict_decl_splitted[decl][page]['destination']
                        for destination in self.dict_decl_splitted[decl][page]['destination']:
                            if not Path(destination).exists():
                                Path(destination).mkdir(parents=True)
                            decl_page_path = destination / str(decl_page_name)
                            with open(decl_page_path, 'wb') as decl_page:
                                # print('in for 4')
                                # page_pdf_writer.write(decl_page)
                                # page_pdf_writer.close()
                                # page_pdf_writer = None
                                writer_saver.write(decl_page)
                                # writer_saver = None
                    except:
                        # print('No need to copy this declaration page.')
                        pass
                # Pop every page that is copied
                self.dict_decl_splitted.pop(decl)
                if (dnumber % 100) == 0:
                    gc.collect()

        gc.collect()

    def merge_all_in_one(self):
        gc.collect()
        self.dict_to_copy = None
        # self.dict_decl_splitted=None
        self.df_general_invoice = None
        self.df_facture = None
        self.df_routage = None
        self.df_decl = None
        self.df_tds = None
        self.df_coo = None
        self.dict_decl_splitted = None
        gc.collect()

        print('Merging all files into one...')
        # pdf_merger=PdfFileMerger(strict=False)
        # writer_merger = PdfWriter()
        # Loop through YGT table, and merge all pdfs in one table
        pdf_counter = 0
        less_than_hundred = True
        self.pdf_writer_is_null = True
        for boq in self.dict_situation:
            print('Merging BOQ folder: ', boq)

            # Create new blank pdf and write boq code on it
            self.boq_folder = self.destination / boq

            if self.boq_folder.exists():
                self.boq_pdf_stem = boq + ".pdf"
                self.boq_pdf_full_name = self.boq_folder / self.boq_pdf_stem
                # save FPDF() class into a variable pdf
                pdf = FPDF()
                # Add a page
                pdf.add_page()

                # set style and size of font that you want in the pdf
                pdf.set_font("Arial", "B", size=40)
                # create a multiple cell
                # First create empty cell to center second cell horizontally
                pdf.multi_cell(0, 80, txt='', border = 0,align='C', fill= False)
                pdf.multi_cell(0, 20, txt=boq, border = 1,align='C', fill= False)
                pdf.output(self.boq_pdf_full_name)
                pdf = None
                # It has to be added once when looping inside boq facture
                # So I keep this boolean here
                self.is_boq_page_added = False
            else:
                self.is_boq_page_added = True
                self.not_found_docs['other'].append('There is no Boq folder named: ' + str(boq) + ' in ' +
                                                    str(self.destination) + 'That is what I can\'t create title page for this boq number.')

            for facture in self.dict_situation[boq]:
                # Sometimes they write space at end of facture, that space gets lost when creating folder
                # So, if path_to_loop cannot be found, then add it to unmerged and keep doing next one
                # try:
                path_to_loop = self.destination / str(boq) / str(facture)
                for pdf in path_to_loop.glob('*.pdf'):
                    less_than_hundred = True
                    if self.pdf_writer_is_null:
                        writer_merger = PdfWriter()
                    self.pdf_writer_is_null = False
                    pdf_counter += 1

                    # Here add boq page as title page for once
                    if not self.is_boq_page_added:
                        reader_merger = PdfReader(self.boq_pdf_full_name)
                        writer_merger.addpages(reader_merger.pages)
                        self.is_boq_page_added = True

                    # Add pdfs inside facture folder into writer_merger
                    try:
                        reader_merger = PdfReader(pdf)
                        writer_merger.addpages(reader_merger.pages)
                        reader_merger = None
                    except:
                        self.not_found_docs['other'].append('This pdf cannot be merged {pdf} {boq}'.format(pdf=pdf, boq=boq))
                    # If there is a memory issue then save part of pdf and continue merging new pdfs
                    if (pdf_counter % 150) == 0 or available_ram(type='percentage') < 10:
                        self.pdf_writer_is_null = True
                        less_than_hundred = False

                        writer_merger.write(
                            self.destination / next_available_name(self.destination / 'Situation All In One.pdf'))
                        writer_merger = None
                        # writer_merger = PdfWriter()
                        gc.collect()
                # except:
                #     print('This folder was not merged {}'.format(str(facture)))
                #     self.not_found_docs['facture'].append(facture)
                #     continue

        if less_than_hundred:
            if pdf_counter > 1:
                writer_merger.write(
                    self.destination / next_available_name(self.destination / 'Situation All In One.pdf'))
                writer_merger = None
                print('Last part also merged and saved successfully!')

        # writer_merger.write(self.destination / next_available_name(self.destination / 'Situation All In One.pdf'))

    def add_prename_to_full_file_name(self, full_path, prename):
        parent = full_path.parent
        # name=full_path.name
        stm = full_path.stem
        sfx = str(full_path.suffix).lower()
        # new_name='{prename}{name}'.format(prename = prename, name = name)
        new_name = '{prename}{stm}{sfx}'.format(
            prename=prename, stm=stm, sfx=sfx)
        prename_added_full_path = parent / new_name
        return prename_added_full_path

    def write_not_founds_to_excel(self):
        self.cannot_be_found_xl = Workbook()
        self.temp_sh = self.cannot_be_found_xl.active
        # self.temp_sh.cell(row=1, column=1).value = 1
        for i, type in enumerate(self.not_found_docs, start=1):
            for j, item in enumerate(self.not_found_docs[type], start=1):
                self.temp_sh.cell(row=j, column=i).value = item

        self.wb_name_to_save = str(next_available_name(
            Path(self.destination) / 'I couldn\'t find these documents.xlsx'))
        self.wb_full_name_to_save = Path(
            self.destination) / self.wb_name_to_save
        try:
            self.cannot_be_found_xl.save(self.wb_full_name_to_save)
        except:
            print('Couldn\'t save excel that contains Not Found Documents.')


if __name__ == '__main__':
    start = time.perf_counter()

    xpath = Path(r'D:\BYTK_Facturation\TBA')
    destination_path = Path.cwd() / 'payment'
    situation_file = Path(
        r'D:\BYTK_Facturation\TBA\Material Sanawy Turkmenbasy Bank-A1 rev05 - again print.xlsx')
    general_invoice_file = Path(
        r'D:\BYTK_Facturation\TBA\GENERAL INVOICE rev25.xlsm')

    test = Payment(source_parent=xpath, destination=destination_path,
                   situation_file=situation_file, general_invoice_file=general_invoice_file)
    test.assign_source_files_to_dataframe()
    test.assign_situation_file_to_dictionary()
    test.assign_general_invoice_to_dataframe()
    test.start_searching()
    test.copy_files()
    test.merge_all_in_one()
    test.write_not_founds_to_excel()

    # print(test.df_general_invoice.head())

    finish = time.perf_counter()
    print('Time spent: ', round(finish-start, 2))
