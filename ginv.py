from pathlib import Path
from openpyxl import Workbook, load_workbook
import os
import re
import threading
import time
import pandas as pd
from multiprocessing import Pool

from helper import *



# ginv = {}
all_files = {}


class AssignGInvoice:
    def __init__(self):
        self.ginv = {}
        self.wb = load_workbook('Book1.xlsx')
        self.ws = self.wb['GENERAL INVOICE']

        self.total_factures = len(self.ws['A'])

        for i in range(3, self.total_factures+1):
            # print(self.ws.cell(row=i, column=1).value)
            self.root_facture = self.ws.cell(row=i, column=1).value
            self.facture_no = self.root_facture
            self.routage_no = self.ws.cell(row=i, column=2).value
            self.decl_no = self.ws.cell(row=i, column=3).value
            self.tds_no = self.ws.cell(row=i, column=4).value
            self.coo_no = self.ws.cell(row=i, column=5).value

            self.ginv[self.root_facture] = {
                'facture': {'no': self.facture_no}}
            self.ginv[self.root_facture].update(
                {'routage': {'no': self.routage_no}})
            self.ginv[self.root_facture].update({'decl': {'no': self.decl_no}})
            self.ginv[self.root_facture].update({'tds': {'no': self.tds_no}})
            self.ginv[self.root_facture].update({'coo': {'no': self.coo_no}})

    def find_files(self):

        self.assign_all_files_into_list()

        self.active_threads = []
        self.threads_full = False
        self.threads_counter=0
        # self.root_factures=[]
        for root_facture in ginv:
            # self.root_factures.append(root_facture)
            self.root_facture=root_facture
            # Loop through dictionary that contains all facture information
            
            # Call thread function
            # try:
            self.active_thread = threading.Thread(target=self.thread_function , daemon=True)
            self.active_thread.start()  # start each thread
            # self.threads_counter += 1    #thread counter
            # self.active_threads.append(self.active_thread)
                

            # except RuntimeError:  # too many throws RuntimeError
            #     self.threads_full = True
            #     pass
            #     os.system('pause')

            # finally:
            #     if self.threads_full:
            #         self.threads_counter=0
            #         for thread in self.active_threads:
            #             thread.join()
            #         self.active_threads = []

            #         try:
            #             self.failed_thread.join()
            #         except:
            #             pass
            #         self.failed_thread = threading.Thread(target=self.thread_function, daemon=True)
            #         self.failed_thread.start()  # start each thread
            #         self.threads_counter += 1  # thread counter
                
            # print(self.threads_counter)

    def thread_function(self):
        for file_type in ginv[self.root_facture]:
                self.file_name_to_search = ginv[self.root_facture][file_type]['no']
                if self.file_name_to_search != None:
                    self.files_of_specific_type=all_files[file_type]
                    # Loop through folder that contains pdf files
                    self.files_matched = []
                    for pdf_file_name in self.files_of_specific_type:
                        self.pdf_file_name=pdf_file_name
                        # Comparison here
                        self.pattern = r'(\D|\b)' + str(self.file_name_to_search) + r'\D.+pdf$'
                        self.temp_match = re.match(self.pattern, self.pdf_file_name)

                        if self.temp_match:
                            self.files_matched.append(self.files_of_specific_type[self.pdf_file_name])

        # print(self.files_matched)
        # print('stop')

    def assign_all_files_into_list(self):
        for file_type in overall_path['source']:
            all_files[file_type] = {}
            for xfile in overall_path['source'][file_type].rglob('*.pdf'):

                all_files[file_type].update({xfile.name: xfile})


    def test(self):
        facture_list=[]
        
        for pdf in overall_path['source']['facture'].glob('*.pdf'):
            facture_list.append(str(pdf))

        self.df=pd.DataFrame(facture_list, columns=['facture_no'])
        #print(df.head())

        self.iter_to_pool=[]
        for root in self.ginv:
            self.iter_to_pool.append(root)

        self.test_active_threads=[]
        self.test_active_thread_objects=[]

        p=Pool()
        result=p.map(self.pool_thread, self.iter_to_pool)
        p.close()
        p.join()

        print(result)
    def pool_thread(self, root):
        # for root in kwargs['ginv']:
        if self.ginv[root]['facture']!="":
            for _ in self.ginv[root]['facture']:
                self.no_to_find=self.ginv[root]['facture']['no']
                self.df_to_search=self.df
                # print('Search results for {}'.format(self.no_to_find))
                self.test_thread_pattern=r'(\D|\b)' + str(self.no_to_find) + r'\D.+pdf$'
                # self.test_thread_match='this is sample result'
                self.test_thread_match=self.df_to_search[self.df_to_search['facture_no'].str.contains(self.test_thread_pattern)==True]
                print('Search results for {}'.format(self.no_to_find), self.test_thread_match)
                return {root:self.test_thread_match}
    
def test_thread(self):
    print('TEST thread')

def caller_thread_search(self, **kwargs):
    # if len(self.test_active_threads)>100:
    #     for thread in self.test_active_threads:
    #         thread.join()
    #     self.test_active_threads=[]
    
    self.test_active_thread_objects.append(ThreadSearch())
    self.test_active_thread_objects[len(self.test_active_thread_objects)-1].Search(no_to_find= ginv[kwargs['root']]['facture']['no'], df_to_search= kwargs['df'])

class ThreadSearch:
    # def __init__(self, *args, **kwargs):
    #     self.no_to_find=kwargs['no_to_find']
    #     self.df_to_search=kwargs['df_to_search']
    
    # def Search(self, **kwargs):
    #     self.no_to_find=kwargs['no_to_find']
    #     self.df_to_search=kwargs['df_to_search']
    #     # print('Search results for {}'.format(self.no_to_find))
    #     self.test_thread_pattern=r'(\D|\b)' + str(self.no_to_find) + '\D.+pdf$'
    #     # self.test_thread_match='this is sample result'
    #     self.test_thread_match=self.df_to_search[self.df_to_search['facture_no'].str.contains(self.test_thread_pattern)==True]
    #     print('Search results for {}'.format(self.no_to_find), self.test_thread_match)
    #     #os.system('pause')
    pass



if __name__ == '__main__':
    start = time.perf_counter()

    test = AssignGInvoice()
    # test.find_files()
    test.assign_all_files_into_list()
    #test.find_files()

    test.test()
    # counter=0
    # try:
    #     for t in test.test_active_threads:
            
    #         t.start()

    #         counter+=1
    #         if counter>600:
    #             os.system('pause')
    # except:
    #     print('Number of total threads {}'.format(len(test.test_active_threads)))
    #     print('Number of started threads {}'.format(counter))
    #     os.system('pause')
    # for t in test.test_active_threads:
    #     t.join()
    # print('Number of total threads {}'.format(len(test.test_active_threads)))
    # print('Number of started threads {}'.format(counter))
    finish=time.perf_counter()

    print('Total time spent: {}'.format(round(finish-start,2)))
