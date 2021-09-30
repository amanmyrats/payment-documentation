from pathlib import Path
import re
import time
import sys

import pandas as pd
from openpyxl import Workbook, load_workbook
import win32com.client as win32
import xlwings as xw
import tempfile
import shutil

from helper import next_available_name
from last_modife import LastModifeFinder


class AnalyzeForDataFrame:
    def __init__(self, **kwargs):
        self.is_facture = False
        if kwargs['type'] == 'df':
            self.wb = kwargs['wb']
            if self.has_annexe_sheet() and self.has_facture_sheet():
                self.is_facture = True

    def has_facture_sheet(self):
        try:
            self.facture_df = pd.read_excel(self.wb, sheet_name='Facture')
            return True
        except:
            # ERROR NOT FACTURE
            return False

    def has_annexe_sheet(self):
        try:
            self.annexe_df = pd.read_excel(self.wb, sheet_name='Annexe')
            return True
        except:
            try:
                self.annexe_df = pd.read_excel(self.wb, sheet_name='Annexe(2)')
                return True
            except:
                # ERROR NOT FACTURE
                return False

    def get_facture_df(self):
        try:
            self.facture_df = pd.read_excel(self.wb, sheet_name='Facture')
            return self.facture_df
        except:
            # ERROR NOT FACTURE
            return None

    def get_annexe_df(self):
        try:
            self.annexe_df = pd.read_excel(self.wb, sheet_name='Annexe')
            return self.annexe_df
        except:
            try:
                self.annexe_df = pd.read_excel(self.wb, sheet_name='Annexe(2)')
                return self.annexe_df
            except:
                # ERROR NOT FACTURE
                return None


class AnalyzeMaterialTable:
    def __init__(self, material_table, result_columns):
        print('Assigning material table to dictionary...')
        self.material_table=load_workbook(material_table)
        self.sh_all_factures=self.material_table['ALL FACTURES']
        self.result_columns=result_columns
        self.material_dict={}
        self.new_factures_to_transfer={}

    def assign_to_dictionary(self):
        self.find_last_row()
        for i in range(2,self.last_row+1):
            # print('assigning: {} - {} '.format(i, self.last_row))
            # self.material_dict={
                #     '001-NCC-TK' : {
                #                         1 : {
                #                                 'warning'   : 'if there is any, warning will be only in order_no:1',
                #                                 'row'       : 121
                #                                 6           : 'cell value of row:121 and column:6',
                #                                 7           : 'cell value of row:121 and column:7',
                #                                 8           : 'cell value of row:121 and column:8',
                #                                 9           : 'cell value of row:121 and column:9',
                #                                 10           : 'cell value of row:121 and column:10',
                #                                 11           : 'cell value of row:121 and column:11',
                #                                 12           : 'cell value of row:121 and column:12',
                #                                 13           : 'cell value of row:121 and column:13',
                #                                 14           : 'cell value of row:121 and column:14',
                #                                 15           : 'cell value of row:121 and column:15',
                #                                 16           : 'cell value of row:121 and column:15',
                #                                 21           : 'cell value of Excel File Name row:121 and column:16'
                #                         }
                #     }
                # }


            row=i
            facture_col=self.result_columns['facture_no']['colno']
            order_no_col=self.result_columns['order_no']['colno']
            facture_no = self.sh_all_factures.cell(row=row, column=facture_col).value

            

            # facture_no key
            try:
                self.material_dict[facture_no]
            except:
                self.material_dict[facture_no]={}

            # order_no key
            order_no= self.sh_all_factures.cell(row=row, column=order_no_col).value
            
            # if facture_no=='1585-TK-NCC' and order_no==2:
            #     print('stop')

            try:
                self.material_dict[facture_no][order_no]
                self.material_dict[facture_no][order_no]['warning']='duplicate_material'
                # self.list_delete_later.append(facture_no)
            except:
                self.material_dict[facture_no][order_no]={}

            # Which row of excel file
            self.material_dict[facture_no][order_no]['row']=row

            # values of one row
            for coltitle in self.result_columns:
                temp_column=self.result_columns[coltitle]['colno']
                temp_data=self.sh_all_factures.cell(row=row, column=temp_column).value
                
                self.material_dict[facture_no][order_no][temp_column]=temp_data

    def find_last_row(self):
        self.last_row=self.sh_all_factures.max_row
        if self.last_row<300000:
            return self.last_row
        else:
            self.last_row = 300000
            return self.last_row

    def check_if_exists(self):
        pass

    def collect_new_facture(self, all_factures_path):
        """
        takes self, all_factures_path
        """
        self.all_factures_path=all_factures_path
        self.facture_excel_names=[]

        for xl in self.all_factures_path.rglob('*.xls*'):
            self.facture_excel_names.append(xl.name)

        
        self.df_facture_names=pd.DataFrame(self.facture_excel_names, columns=['excel_names'])
        # print('Type: ', type(self.df_facture_names))

        self.only_new_facture_names=[]
        self.modife_finder_after_regex=LastModifeFinder(parent_path=self.all_factures_path)
        
        xxx=1
        self.facture_names_not_in_material_table=list(self.facture_excel_names)
        self.facture_names_modife=[]
        self.combined_facture_names_to_transfer=[]
        self.list_combined_path_of_factures_to_transfer=[]

        # Loop through facture numbers of material table
        for i, facture_no in enumerate(self.material_dict, start=1):
            self.finder_pattern1=r'.*?(\D|\b)' + str(facture_no) + r'\D.+xl.{1,2}$'
            self.finder_pattern2=r'.*?[^-]' + str(facture_no) + r'\D.+xl.{1,2}$'
            self.temp_series_search_result=self.df_facture_names[self.df_facture_names['excel_names'].str.match(re.escape(self.finder_pattern1), case=False, na=False)]
            self.temp_series_search_result=self.temp_series_search_result[self.temp_series_search_result['excel_names'].str.match(re.escape(self.finder_pattern2), case=False, na=False)]
            
            if len(self.temp_series_search_result.index)==0:
                continue

            self.modife_finder_after_regex.find_last_modife(facture_no=facture_no, file_name_list=self.temp_series_search_result['excel_names'].tolist())

            if self.modife_finder_after_regex.last_modife=='none':
                continue 

            # Match excel file name of material table and all_facture_folder
            # if '001-TK-NCC.xlsm' == '001-TK-NCC modife 1.xlsm':
            if not self.material_dict[facture_no][list(self.material_dict[facture_no])[0]][self.result_columns['excel_file_name']['colno']]==self.modife_finder_after_regex.last_modife:
                # self.new_factures_to_transfer[facture_no]=Path(self.all_factures_path / self.modife_finder_after_regex.last_modife)
                self.facture_names_modife.append(self.modife_finder_after_regex.last_modife)
                print(xxx, ' out of ', i,' have new modife.')
                xxx+=1

            # Remove matched excel file names from initial names
            for xl_to_remove in self.temp_series_search_result['excel_names'].tolist():
                try:
                    self.facture_names_not_in_material_table.remove(xl_to_remove)
                except:
                    # print('ERROR: {} is not in list'.format(xl_to_remove))
                    pass

        print('\n Modife factures to transfer ', len(self.new_factures_to_transfer))
        # print(self.new_factures_to_transfer)
        print('\n New factures to transfer ', len(self.facture_names_not_in_material_table))


        # Combine modife factures and new facture into new list self.list_combined_path_of_factures_to_transfer
        self.facture_names_modife.extend(self.facture_names_not_in_material_table)
        self.combined_facture_names_to_transfer=list(self.facture_names_modife)
        print('\n Combined factures to transfer ', len(self.combined_facture_names_to_transfer))

        # Convert combined file names into file path
        for combined in self.combined_facture_names_to_transfer:
            self.list_combined_path_of_factures_to_transfer.append(Path(self.all_factures_path / combined))
        
    def test(self):
        # for facture in self.material_dict:
        #     for order in self.material_dict[facture]:
        #         for col in self.material_dict[facture][order]:
        #             print(facture , ' - ', order ,' - ',col, ' - ', self.material_dict[facture][order][col] )
        
        print('Length of self.material_dict: ', len(self.material_dict))
        print('Here is the test of AnalyzeMaterialTable')


class InfoFetcher:
    def __init__(self, material_table, all_factures_path, **kwargs):
        self.result_next_row = 2
        self.material_table=material_table
        self.all_factures_path=all_factures_path
        self.create_dictionaries()
        self.assign_ext_excel_to_dict()
        self.start_column_of_material_table = 12
        self.last_column_of_material_table = 30

    def create_dictionaries(self):
        self.facture_cells = {
            'facture_no':               {'cell': 'F2', 'row': 0, 'column': 5},
            'hg':                       {'cell': 'C8', 'row': 6, 'column': 2,'extra':self.hg_works},
            'routage':                  {'cell': 'C32', 'row': 30, 'column': 2,'extra':self.routage_works}, 
            'imp_bp':                   {'cell': 'C36', 'row': 34, 'column': 2,'extra':self.imp_bp_works},
            'zawod_baha':               {'cell': 'F50', 'row': 48, 'column': 5},
            'asgabada_gelen_baha':      {'cell': 'F52', 'row': 50, 'column': 5}
        }

        self.annexe_columns = {
            'facture_date':     {'colno': 12, 'colletter': 'L', 'colname': 'Date', 
                                'result_key':('facture_date',)},
            'designations':     {'colno': 14, 'colletter': 'N', 'colname': 'Designations', 
                                'result_key':('designations_fr', 'designations_ru', 'designations_tm',),
                                'extra':self.split_designation},
            'fournisseur':      {'colno': 15, 'colletter': 'O', 'colname': 'Fournisseur', 
                                'result_key':('fournisseur',),
                                'extra':self.company_works},
            'prod_date':        {'colno': 16, 'colletter': 'P', 'colname': 'Année ', 
                                'result_key':('prod_date',),
                                'extra':self.company_works},
            'pays':             {'colno': 17, 'colletter': 'Q', 'colname': 'Pays ', 
                                'result_key':('pays',),
                                'extra':self.country_works},
            'unit':             {'colno': 18, 'colletter': 'R', 'colname': 'Unité', 
                                'result_key':('unit',),
                                'extra':self.unit_works},
            'qt':               {'colno': 19, 'colletter': 'S', 'colname': 'Qté', 
                                'result_key':('qt',)},
            'monnaie':          {'colno': 20, 'colletter': 'T', 'colname': 'Monnaie', 
                                'result_key':('monnaie',),
                                'extra':self.monnaie_works},
            'pu':               {'colno': 21, 'colletter': 'U', 'colname': 'PU', 
                                'result_key':('pu',)},
            'prix_total':       {'colno': 22, 'colletter': 'V', 'colname': 'Prix Total', 
                                'result_key':('prix_total',)}
        }

        self.result_columns = {
            'order_no':         {'colno': 2, 'colletter': 'B', 'current_value': ''},
            
            # Facture sheet
            'hg':                       {'colno': 12, 'colletter': 'L', 'current_value': ''},
            'facture_no':               {'colno': 14, 'colletter': 'N', 'current_value': ''},
            'routage':                  {'colno': 26, 'colletter': 'Z', 'current_value': ''},
            'imp_bp':                   {'colno': 27, 'colletter': 'AA', 'current_value': ''},
            'zawod_baha':               {'colno': 28, 'colletter': 'AB', 'current_value': ''},
            'asgabada_gelen_baha':      {'colno': 29, 'colletter': 'AC', 'current_value': ''},

            # Annexe sheet
            
            'facture_date':     {'colno': 13, 'colletter': 'M', 'current_value': ''},
            'designations_fr':  {'colno': 15, 'colletter': 'O', 'current_value': ''},
            'designations_ru':  {'colno': 16, 'colletter': 'P', 'current_value': ''},
            'designations_tm':  {'colno': 17, 'colletter': 'Q', 'current_value': ''},
            'fournisseur':      {'colno': 18, 'colletter': 'R', 'current_value': ''},
            'prod_date':        {'colno': 19, 'colletter': 'S', 'current_value': ''},
            'pays':             {'colno': 20, 'colletter': 'T', 'current_value': ''},
            'unit':             {'colno': 21, 'colletter': 'U', 'current_value': ''},
            'qt':               {'colno': 22, 'colletter': 'V', 'current_value': ''},
            'monnaie':          {'colno': 23, 'colletter': 'W', 'current_value': ''},
            'pu':               {'colno': 24, 'colletter': 'X', 'current_value': ''},
            'prix_total':       {'colno': 25, 'colletter': 'Y', 'current_value': ''},
            

            # Others
            'excel_file_name': {'colno': 30, 'colletter': 'AD', 'current_value': ''}
        }

        self.to_transfer={}

    def assign_ext_excel_to_dict(self):

        # Dictionaries of company, country and unit standards are assigned here
        self.external_files_path=Path.cwd() / 'external_files'
        self.naming_standards_excel=self.external_files_path / 'naming_standards.xlsx'
        self.sheet_names_for_naming_standards=('company_standard','country_standard','unit_standard')
        self.dict_stds={
            'company_standard':{},
            'country_standard':{},
            'unit_standard':{}
        }
        for sh_name in self.sheet_names_for_naming_standards:
            # print('excel name: {}'.format(self.naming_standards_excel))
            # print('sheet name: {}'.format(sh_name))
            try:
                self.temp_df=pd.read_excel(self.naming_standards_excel, sheet_name=sh_name)
                print('Assigning external info ', sh_name)
                for index, xrow in self.temp_df.iterrows():
                    self.current_trash=str(xrow['trash'])
                    self.current_corrected=str(xrow['corrected'])
                    try:
                        # print(type(xrow['trash']))
                        if not pd.isnull(self.current_trash) and not self.current_trash.isspace():
                            # print('xrow[trash]: ', xrow['trash'], ' corrected: ', xrow['corrected'])
                            self.dict_stds[sh_name][self.current_trash]=self.current_corrected
                            pass
                    except:
                        print('Error in sheet:{}'.format(sh_name) )
                        pass
            except:
                print('Could not assign sheet {} into dataframe.'.format(sh_name))


        # TEST
        self.material_dict={}
        self.test_material=AnalyzeMaterialTable(material_table=self.material_table, result_columns=self.result_columns)
        print('created analyzematerialtable')
        self.test_material.check_if_exists()
        print('checked if extists')
        self.test_material.assign_to_dictionary()
        print('assigned to dictinoary')

        # TEST
        self.test_material.collect_new_facture(all_factures_path=self.all_factures_path)
        print('collected new factures')

        self.material_dict=self.test_material.material_dict
        print('assigned material_dict')
        self.new_factures_to_transfer=self.test_material.new_factures_to_transfer
        print('assigned new factures to transfer')
        self.list_combined_path_of_factures_to_transfer=self.test_material.list_combined_path_of_factures_to_transfer
        print('assigned combined path of factures')
        print('Length to_transfer before: ', len(self.list_combined_path_of_factures_to_transfer))

    def create_result_wb_for_result(self):
        print('I will create new workbook for result!')
        self.result_wb = Workbook()
        self.result_sh = self.result_wb.active
        self.result_next_row=2
        self.is_win32=False
        
    def eat_facture(self, df_facture, df_annexe, current_excel_file_name):
        self.df_facture = df_facture
        self.df_annexe = df_annexe
        self.current_excel_file_name = current_excel_file_name
        self.total_material = self.material_count()
        # All rows of dataframe that has material
        self.df_rows = self.df_annexe.loc[pd.to_numeric(self.df_annexe['Qté'], errors='coerce') > 0, ['Qté']]
        self.list_index_df_rows=list(self.df_rows.index.values)

        for i, current_df_row in enumerate(self.df_rows.iterrows(), start=1):
            self.current_current_df_row=current_df_row
            self.result_next_row += 1
            self.is_last_material=False
            
            # Write order number and excel file name into result workbook
            if not self.is_win32:
                self.write_to_openpyxl(order_no=i, input_type='order_filename')
            else:
                self.write_to_win32(order_no=i, input_type='order_filename')
                
            # HOW MANY ROW TILL NEXT MATERIAL
            if i==len(self.list_index_df_rows):
                # If it is last material
                self.is_last_material=True
                self.how_many_rows_to_next_material=400
            else:
                self.how_many_rows_to_next_material=self.list_index_df_rows[i]-self.list_index_df_rows[i-1]

            # Loop through FACTURE sheet
            for facture_cell in self.facture_cells:
                self.current_cell=facture_cell 
                try:
                    # Find
                    self.result_columns[self.current_cell]['current_value'] = \
                        self.df_facture.iloc[self.facture_cells[facture_cell]['row'], self.facture_cells[self.current_cell]['column']]
                    self.facture_no=self.df_facture.iloc[self.facture_cells['facture_no']['row'], self.facture_cells['facture_no']['column']]
                    # Call extra function if has one
                    try:
                        self.facture_cells[facture_cell]['extra'](order_no=i)
                    except:
                        pass

                    # Write it to result sheet
                    if not self.is_win32:
                        self.write_to_openpyxl(order_no=i, input_type='facture')
                    else:
                        self.write_to_win32(input_type='facture')
                except:
                    self.result_columns[facture_cell]['current_value'] =''

            # Loop through ANNEXE sheet
            for annexe_column in self.annexe_columns:
                self.current_annexe_column=annexe_column
                self.current_key_of_annexe=self.annexe_columns[self.current_annexe_column]
                self.pays_found=False
                # First empty all old current values of dictionary
                try:
                    self.result_columns[self.current_key_of_annexe['result_key'][0]]['current_value'] =''
                    self.result_columns[self.current_key_of_annexe['result_key'][1]]['current_value'] =''
                    self.result_columns[self.current_key_of_annexe['result_key'][2]]['current_value'] =''
                except:
                    pass

                try:
                    # Find
                    self.result_columns[self.current_key_of_annexe['result_key'][0]]['current_value'] = \
                        self.df_annexe.loc[current_df_row[0],self.current_key_of_annexe['colname']]
                    # Call extra function if has one
                    try:
                        self.annexe_columns[annexe_column]['extra']()
                        # print('Extra function executed successfully for', annexe_column)
                    except:
                        # print('There is no extra function for', annexe_column)
                        pass

                    # Write it to result sheet
                    # All including designation
                    if not self.is_win32:
                         self.write_to_openpyxl(order_no=i, input_type='annexe')
                    else:
                        self.write_to_win32(input_type='annexe')

                    self.pays_found=True
                except:
                    if not self.pays_found:
                        self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value'] = ''

    def transfer_to_material_table(self, **kwargs):
        
        self.modife_matcher=LastModifeFinder(parent_path=self.all_factures_path)

        # Pop factures that are open file's temp file (excel file that has ~$ at the beginning)
        self.open_factures_to_pop=[]
        for temp_open_facture in self.to_transfer:
            if str(temp_open_facture).lower().__contains__('~$'):
                self.open_factures_to_pop.append(temp_open_facture)

        for temp_facture_to_pop in self.open_factures_to_pop:
            try:
                self.to_transfer.pop(temp_facture_to_pop)
            except:
                # print('Could not find open facture\'s temp file to pop')
                pass

        # Pop preparation of same factures already in material table with same version
        # self.set_to_pop=[]
        self.set_to_pop=set()
        for new_facture in self.to_transfer:
            try:
                self.current_facture=self.to_transfer[new_facture][1][self.result_columns['facture_no']['colno']]
                # If new facture does not exist in material table, then exception will happen
                # It means that we have to add that to material table
                self.old_file_name=self.material_dict[self.current_facture][[*self.material_dict[self.current_facture]][0]][self.result_columns['excel_file_name']['colno']]
                self.new_file_name=new_facture
                
                # self.material_dict[self.current_facture]
                
                # handle same file names
                # old_file_name is the file name written in the material table
                if self.old_file_name==self.new_file_name:
                    self.set_to_pop.add(self.new_file_name)
                    continue
                else:
                    # Decide which one is last modife
                    if str(new_facture).__contains__('test'):
                        print('stop')
                    
                    # Find Last Modife
                    try:
                        self.modife_matcher.find_last_modife(facture_no = self.current_facture, file_name_list=[self.old_file_name, self.new_file_name])
                        self.temp_last_modife=self.modife_matcher.last_modife

                        # # TEST 
                        # if self.temp_last_modife=='none':
                        #     print('self.temp_last_modife is none')

                    except:
                        print('Exception happened when matching: ', new_facture)

                    # If last modife is equal with old one, then pop it from to_transfer dictionary
                    if self.temp_last_modife==self.old_file_name:
                        self.set_to_pop.add(self.new_file_name)
                        continue
                    elif self.temp_last_modife=='none':
                        self.set_to_pop.add(self.new_file_name)
                        continue
                    else:
                        # If new facture is last modife, then add it to bottom
                        # If total number of materials in facture is not same
                        # Add it at the end of material table, and write warning note, and make different background color.
                        # If ordering is not proper in material table then add it to bottoma and add warning
                        self.temp_material_count_old = len(self.material_dict[self.current_facture])
                        self.temp_material_count_new = len(self.to_transfer[new_facture])
                        
                        # Convert all rows of current facture no into list(list will contain row numbers of current facture)
                        self.temp_rows_list=[]
                        for temp_order_no in self.material_dict[self.current_facture]:
                            # for temp_col in self.material_dict[self.current_facture][temp_order_no]:
                            self.temp_row_no=self.material_dict[self.current_facture][temp_order_no]['row']
                            self.temp_rows_list.append(self.temp_row_no)

                        # Convert all orderno of current facture no into list(list will contain order numbers of current facture)
                        self.temp_orderno_list=[]
                        for temp_order_no2 in self.material_dict[self.current_facture]:
                            self.temp_orderno_list.append(temp_order_no2)
                        
                        # Get warning message from material table if there is any
                        # This warning message has been written already when material_dict is assigned, if there is any
                        # Warning message is about duplicate order numbers in material table
                        try:
                            # self.warning_duplicate_material=self.material_dict[self.current_facture][[*self.material_dict[self.current_facture]][0]]['warning']
                            # self.warning_duplicate_material=self.material_dict[self.current_facture][1]['warning']
                            for temp_order_for_warning in self.material_dict[self.current_facture]:
                                try: 
                                    self.material_dict[self.current_facture][temp_order_for_warning]['warning']
                                    self.temp_warning_getter=self.material_dict[self.current_facture][temp_order_for_warning]['warning']
                                    if len(self.temp_warning_getter)>0:
                                        self.warning_duplicate_material=self.temp_warning_getter
                                        print('Warning message caught: ', self.temp_warning_getter)
                                        break
                                except:
                                    self.warning_duplicate_material=''
                                    # print('No warning message assigned during material_dict creation.')
                        except:
                            print('Error: Could not get warning message')
                            self.warning_duplicate_material=''

                        # Classify new factures into inline and bottom
                        if not self.temp_material_count_new == self.temp_material_count_old:
                            self.to_transfer[new_facture][1]['warning'] = 'Duplicate! Number of total materials are not same.'
                            self.to_transfer[new_facture][1]['where'] = 'bottom'
                        # if order no is duplicate
                        elif self.warning_duplicate_material=='duplicate_material':
                            self.to_transfer[new_facture][1]['warning'] = 'Duplicate, order no is duplicate please check.'
                            self.to_transfer[new_facture][1]['where'] = 'bottom'
                        # Feed row list of current facture into funtion, and function will know if it is queue or not
                        elif self.temp_material_count_new == self.temp_material_count_old and self.is_rows_queue(number_list=self.temp_rows_list) and self.is_ordernos_queue(number_list=self.temp_orderno_list):
                            self.to_transfer[new_facture][1]['where'] = 'inline'
                        # Decide if materials of old facture is placed as a queue
                        # If it is not a queue then add it to bottom with warning
                        else:
                            self.to_transfer[new_facture][1]['warning'] = 'Duplicate! Existing materials are not queued.'
                            self.to_transfer[new_facture][1]['where'] = 'bottom'

            except:
                # if exception happened, it means that we need that material, do nothing for now
                print('It does not exist in material table, will be added to bottom: ', new_facture)
                self.to_transfer[new_facture][1]['where']='bottom'
                pass
        
        # Pop factures those are already in Material Table
        for fname in self.set_to_pop:
            self.to_transfer.pop(fname)

        # Pop materials those are duplicate inside to_transfer
        self.set_to_pop_after_duplicate_check=set()
        for possible_duplicate_file_name in self.to_transfer:
            self.possible_duplicate_facture_no=self.to_transfer[possible_duplicate_file_name][1][self.result_columns['facture_no']['colno']]
            for other_file_name in self.to_transfer:
                if possible_duplicate_file_name==other_file_name:
                    continue

                self.temp_facture_no_for_dupicate_check=self.to_transfer[other_file_name][1][self.result_columns['facture_no']['colno']]
                if self.possible_duplicate_facture_no==self.temp_facture_no_for_dupicate_check:
                    # Check which one is last modife
                    self.modife_matcher.find_last_modife(facture_no = self.possible_duplicate_facture_no, file_name_list=[possible_duplicate_file_name, other_file_name])
                    self.last_modife_for_duplicate_check=self.modife_matcher.last_modife

                    if self.last_modife_for_duplicate_check==possible_duplicate_file_name:
                        self.set_to_pop_after_duplicate_check.add(other_file_name)
        
        for duplicate in self.set_to_pop_after_duplicate_check:
            # try:
            self.to_transfer.pop(duplicate)
            # except:
            #     print('Did not find {} to pop from duplicates.'.format(duplicate))
            #     pass

        print('Length to_transfer after: ', len(self.to_transfer))
        print('Length of copied material_dict: ', len(self.material_dict))
        
        # TRANSFERING STARTS HERE
        try:
            self.wb=xw.Book(str(self.material_table), update_links=False, notify=False)
        except:
            try:
                if Path(Path(tempfile.gettempdir()) / 'gen_py').exists():
                    shutil.rmtree(Path(Path(tempfile.gettempdir()) / 'gen_py'))
                    print('I deleted gen_py :)')
                self.wb=xw.Book(str(self.material_table), update_links=False, notify=False)
            except:
                print('Error when trying to open Material Table, failed to close gen_py.')
                sys.exit(0)

        self.wb.app.calculation='manual'
        self.wb.app.visible=False
        self.wb.app.dispay_alerts=False
        self.wb.app.screen_updating=False

        self.sheet=self.wb.sheets['ALL FACTURES']

        self.test_material.find_last_row()
        self.next_row_of_pywin32=self.test_material.last_row + 10
        self.is_inline=False
        self.list_to_add=[]
        
        for new_facture_to_add in self.to_transfer:
            # print('adding to pywin32: ', new_facture_to_add)
            try:
                if self.to_transfer[new_facture_to_add][1]['where']=='inline':
                    self.is_inline=True
                elif self.to_transfer[new_facture_to_add][1]['where']=='bottom':
                    self.is_inline=False
            except:
                pass

            # INLINE
            if self.is_inline:
                self.list_inline_to_add=[]
                self.current_inline_facture_no=self.to_transfer[new_facture_to_add][1][self.result_columns['facture_no']['colno']]
                self.first_row_of_current_facture=self.material_dict[self.current_inline_facture_no][1]['row']
                self.current_row_to_color=self.first_row_of_current_facture
                for inlineorderno in self.to_transfer[new_facture_to_add]:
                    self.list_inline_row=[]
                    for j in range(self.start_column_of_material_table, self.last_column_of_material_table + 1):
                        try:
                            self.temp_inline_val = self.to_transfer[new_facture_to_add][inlineorderno][j]
                            self.list_inline_row.append(self.temp_inline_val)
                        except:
                            self.list_inline_row.append('')

                    self.list_inline_to_add.append(self.list_inline_row)
                    self.sheet.range(self.current_row_to_color, self.result_columns['hg']['colno']-1).color=(255,0,0)
                    self.current_row_to_color+=1
                    
                self.sheet.range(self.first_row_of_current_facture, self.result_columns['hg']['colno']).value=self.list_inline_to_add
                

            else:
                # BOTTOM
                self.warning_message=''
                for orderno in self.to_transfer[new_facture_to_add]:
                    self.list_row=[]
                    for i in range(1, self.last_column_of_material_table + 1):
                        try:
                            # self.temp_col = i
                            # It will throw exception when i==1, because only factures with warning has 1 as column no
                            self.temp_val = self.to_transfer[new_facture_to_add][orderno][i]
                            self.list_row.append(self.temp_val)
                        except:
                            if i==1:
                                try:
                                    self.temp_val = self.to_transfer[new_facture_to_add][orderno]['warning']
                                    self.list_row.append(self.temp_val)
                                    self.warning_message=self.temp_val
                                except:
                                    # Add warning message to every 1st column of rows, if there is no warning message then it will be empty.
                                    self.list_row.append(self.warning_message)
                            else:
                                self.list_row.append('')
                                
                    self.list_to_add.append(self.list_row)
                print('Adding new facture: ', new_facture_to_add)
        
                # Add all bottom factures at once
                self.sheet.range(self.next_row_of_pywin32, 1).value=self.list_to_add
        
        self.fstem=Path(self.material_table).stem
        self.fsuffix=Path(self.material_table).suffix
        self.fname=self.fstem + ' - Updated' + self.fsuffix

        self.full_path=Path(self.material_table).parent / self.fname
        self.full_path=self.full_path.parent / next_available_name(full_path=self.full_path)

        print('Saving: ', self.full_path)
        self.wb.app.calculation='automatic'
        self.wb.app.visible=True
        self.wb.app.dispay_alerts=True
        self.wb.app.screen_updating=True
        try:
            self.wb.save(self.full_path)
            print('Saved: ', self.full_path)
        except:
            msg = 'You cannot save this workbook with the same name as another open workbook or add-in. Choose a different name, or close the other workbook or add-in before saving.'
            print(msg)
            import tkinter
            from tkinter import messagebox

            xroot = tkinter.Tk()
            xroot.withdraw()

            messagebox.showinfo('Cannot Save Excel', msg)
    
    def save_result_wb_after_done(self):
        # if self.material_table=='new':
        self.result_wb.save('test.xlsx')
        # else:
        #     try:
        #         self.nextname=0
        #         self.full_path=Path.cwd() / 'test material table.xlsm'
        #         while self.full_path.exists():
        #             self.nextname+=1
        #             self.newname='test material table - ' + str(self.nextname) + '.xlsm'
        #             self.full_path=Path.cwd() / self.newname

        #         self.material_table_wb.SaveAs(str(self.full_path))
        #         print('saved ', self.full_path)
        #         self.excel.Visible=True
        #     except:
        #         print('Error happened when saving pywin32')

    def write_to_openpyxl(self, **kwargs):
        if kwargs['input_type']=='order_filename':
            self.result_sh.cell(
                                row=self.result_next_row,
                                column=self.result_columns['order_no']['colno'],
                                value=kwargs['order_no'])
            self.result_sh.cell(
                        row=self.result_next_row,
                        column=self.result_columns['excel_file_name']['colno'],
                        value=self.current_excel_file_name)
            try:
                self.to_transfer[self.current_excel_file_name]
            except:
                self.to_transfer[self.current_excel_file_name]={}
                # self.to_transfer[self.current_excel_file_name][1000]='test'

            try:
                self.to_transfer[self.current_excel_file_name][kwargs['order_no']]
            except:
                self.to_transfer[self.current_excel_file_name][kwargs['order_no']]={}
                # self.to_transfer[self.current_excel_file_name]['facture_no']=self.facture_no
                # self.to_transfer[self.current_excel_file_name]['total_material']=self.total_material

            self.to_transfer[self.current_excel_file_name][kwargs['order_no']][self.result_columns['order_no']['colno']]= kwargs['order_no']
            self.to_transfer[self.current_excel_file_name][kwargs['order_no']][self.result_columns['excel_file_name']['colno']]= self.current_excel_file_name

        elif kwargs['input_type']=='facture':
            self.result_sh.cell(
                            row=self.result_next_row,
                            column=self.result_columns[self.current_cell]['colno'],
                            value=self.result_columns[self.current_cell]['current_value'])
            
            self.to_transfer[self.current_excel_file_name][kwargs['order_no']][self.result_columns[self.current_cell]['colno']]=self.result_columns[self.current_cell]['current_value']
 
        elif kwargs['input_type']=='annexe':
            self.result_sh.cell(
                row=self.result_next_row,
                column=self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['colno'],
                value=self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value'])

            # try:
            #     self.to_transfer[self.current_excel_file_name][kwargs['order_no']][self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['colno']]
            # except:
            #     self.to_transfer[self.current_excel_file_name][kwargs['order_no']][self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['colno']]={}
            # Equivalent to
            # self.to_transfer['001-TK-NCC.xlsx'][1]['fournisseur']['colno'] = 'SCHNEIDER'
            # self.to_transfer['001-TK-NCC.xlsx'][1][18] = 'SCHNEIDER'
            self.to_transfer[self.current_excel_file_name][kwargs['order_no']][self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['colno']]=self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value']
            # print(self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value'])

            try:
                # Only designation
                self.result_sh.cell(
                    row=self.result_next_row,
                    column=self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][1]]['colno'],
                    value=self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][1]]['current_value'])
                self.result_sh.cell(
                    row=self.result_next_row,
                    column=self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][2]]['colno'],
                    value=self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][2]]['current_value'])

                # Equivalent to
                # self.to_transfer['001-TK-NCC.xlsx'][1]['designations_ru']['colno'] = 'MERMER'
                # self.to_transfer['001-TK-NCC.xlsx'][1]['designations_tm']['colno'] = 'MERMER'
                # self.to_transfer['001-TK-NCC.xlsx'][1][16] = 'МРАМОР'
                # self.to_transfer['001-TK-NCC.xlsx'][1][17] = 'MERMER'
                self.to_transfer[self.current_excel_file_name][kwargs['order_no']][self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][1]]['colno']]=self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][1]]['current_value']
                self.to_transfer[self.current_excel_file_name][kwargs['order_no']][self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][2]]['colno']]=self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][2]]['current_value']
            except:
                pass

    def write_to_win32(self, **kwargs):
        if kwargs['input_type']=='order_filename':
            self.material_table.Cells(self.result_next_row, self.result_columns['order_no']['colno']).Value=kwargs['order_no']
            self.material_table.Cells(self.result_next_row, self.result_columns['excel_file_name']['colno']).Value=self.current_excel_file_name
        elif kwargs['input_type']=='facture':
            self.material_table.Cells(self.result_next_row, 
                            self.result_columns[self.current_cell]['colno']).Value=self.result_columns[self.current_cell]['current_value']
        elif kwargs['input_type']=='annexe':
            self.material_table.Cells(self.result_next_row, 
                self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['colno']
                ).Value=self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value']

            try:
                self.material_table.Cells(self.result_next_row, 
                    self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][1]]['colno']
                    ).Value=self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][1]]['current_value']

                self.material_table.Cells(self.result_next_row, 
                    self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][2]]['colno']
                    ).Value=self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][2]]['current_value']

            except:
                pass

    def material_count(self):
        counter = self.df_annexe.loc[pd.to_numeric(
            self.df_annexe['Qté'], errors='coerce') > 0, ['Qté']].shape[0]
        # print("This is the value that material_count returns: ", counter)
        return counter

    def annexe_column_names_test(self, df_annexe, current_excel_file_name):
        self.test_df_annexe = df_annexe
        self.current_excel_file_name = current_excel_file_name
        self.result_next_row += 1
        self.test_columns = self.test_df_annexe.columns

        self.result_sh.cell(
            row=self.result_next_row,
            column=1,
            value=self.current_excel_file_name
        )

        for i, column in enumerate(self.test_columns, start=2):

            self.result_sh.cell(
                row=self.result_next_row,
                column=i,
                value=column
            )

    def split_designation(self):
        self.is_after_russian=False
        # print('Now it will loop {} times till end of material names'.format(self.how_many_rows_to_next_material-1))
        for name_counter in range(1,self.how_many_rows_to_next_material):

            self.is_fr_or_en=False
            self.is_ru=False
            self.is_tm=False
            self.current_material_description=self.df_annexe.loc[self.current_current_df_row[0]+name_counter,self.annexe_columns[self.current_annexe_column]['colname']]
            if pd.isnull(self.current_material_description):
                continue
            elif str(self.current_material_description).lower()=='total' \
                or str(self.current_material_description).lower()=='total page precedente' \
                    or str(self.current_material_description).lower()=='total general':
                break

            
            self.material_name_to_evaluate=self.current_material_description
            self.evaulate_material_name()
            if self.is_fr_or_en:
                self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value'] += ' ' + \
                                self.current_material_description
                self.is_fr_or_en=False
            elif self.is_ru:
                self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][1]]['current_value'] += \
                                self.current_material_description + ' '
                self.is_ru=False
            elif self.is_tm:
                self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][2]]['current_value'] += \
                                self.current_material_description + ' '
                self.is_tm=False
        
    def evaulate_material_name(self):
        self.all_russian_chars="АаБбВвГгДдЕеЁёЖжЗзИиЙйКкЛлМмНнОоПпСсТтУуФфХхЦцЧчШшЩщЪъЫыЬьЭэЮюЯя"
        try:
            if bool(re.search('[а-яА-Я]', self.material_name_to_evaluate)):
                self.len_of_description=len(self.material_name_to_evaluate)
                self.total_russian_match=0
                for x_char in self.material_name_to_evaluate:
                    if x_char in self.all_russian_chars:
                        self.total_russian_match+=1
                
                self.russian_matched_percentage=self.total_russian_match / self.len_of_description * 100
                if int(self.russian_matched_percentage)>3:
                    self.is_ru=True
                    self.is_after_russian=True
                elif self.is_after_russian:
                    self.is_tm=True
                else:
                    self.is_fr_or_en=True

            elif self.is_after_russian:
                self.is_tm=True
            else:
                self.is_fr_or_en=True
        except:
            print('Error material language check: ', self.material_name_to_evaluate)

    def company_works(self):
        # If any of key from the company_dict contains company name then assign it as current value
        try:
            for trash in self.dict_stds['company_standard'].keys():
                if self.dict_stds['company_standard'][trash]=='-':
                    continue
                self.country_tm=self.df_annexe.loc[self.current_current_df_row[0],self.annexe_columns[self.current_annexe_column]['colname']]
                # print('Inside nested dict')
                # print('unit_standard key: ', trash)
                # print('I will search: ', self.country_tm)
                if trash.lower().__contains__(str(self.country_tm).lower()):
                    # print('Match found')
                    self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value']=self.dict_stds['company_standard'][trash].upper()
                    break
        except:
            print('Error in def company_works(self):')

    def country_works(self):
        # print('Inside country_works function.')
        # self.french_country_name = self.df_annexe.loc[self.current_current_df_row[0],self.annexe_columns[self.current_annexe_column]['colname']]
        # if 'diver' in str(self.french_country_name).lower():
        #     self.existing_country_name = self.material_dict[]
        # else:
        for country_counter in range(1,3):
            self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value'] += ' / ' + \
                            self.df_annexe.loc[self.current_current_df_row[0]+country_counter,self.annexe_columns[self.current_annexe_column]['colname']]
        
        # If any of key from the country_dict contains Turkmen version of country then assign it as current value
        try:
            for trash in self.dict_stds['country_standard'].keys():
                if self.dict_stds['country_standard'][trash]=='-':
                    continue
                self.country_tm=self.df_annexe.loc[self.current_current_df_row[0]+2,self.annexe_columns[self.current_annexe_column]['colname']]
                # print('Inside nested dict')
                # print('unit_standard key: ', trash)
                # print('I will search: ', self.country_tm)
                if trash.lower().__contains__(str(self.country_tm).lower()):
                    # print('Match found')
                    self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value']=self.dict_stds['country_standard'][trash].upper()
                    break
        except:
            print('Error in def country_works(self):')

    def unit_works(self):
        for unit_counter in range(1,3):
            self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value'] += ' / ' + \
                            self.df_annexe.loc[self.current_current_df_row[0]+unit_counter,self.annexe_columns[self.current_annexe_column]['colname']]

        # If any of key from the self.dict_stds['unit_standard'] contains Turkmen version of unit then assign it as current value
        try:
            for trash in self.dict_stds['unit_standard'].keys():
                self.unit_tm=self.df_annexe.loc[self.current_current_df_row[0]+2,self.annexe_columns[self.current_annexe_column]['colname']]
                # print('Inside nested dict')
                # print('unit_standard key: ', trash)
                # print('I will search: ', self.unit_tm)
                if trash.lower().__contains__(str(self.unit_tm).lower()):
                    # print('Match found')
                    self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value']=self.dict_stds['unit_standard'][trash].lower()
                    break
        except:
            print('Error in def unit_works(self):')

    def monnaie_works(self):
        for monnaie_counter in range(1,3):
            self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value'] += ' / ' + \
                            self.df_annexe.loc[self.current_current_df_row[0]+monnaie_counter,self.annexe_columns[self.current_annexe_column]['colname']]

        if str(self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value']).lower().__contains__('dollar') \
            or str(self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value']).upper().__contains__('ДОЛЛАР'):
            self.result_columns[self.annexe_columns[self.current_annexe_column]['result_key'][0]]['current_value']='$'

    def routage_works(self, **kwargs):
        self.unmodified_routage=self.result_columns[self.current_cell]['current_value']
        while not str(self.unmodified_routage)[0].isnumeric():
            # print('Inside while in routage_works')
            self.unmodified_routage=self.unmodified_routage[1:]
        
        self.result_columns[self.current_cell]['current_value']=self.unmodified_routage
        # print("Inside routage works")
        pass
    
    def hg_works(self, **kwargs):
        self.modified_hg = ''
        self.unmodified_hg=self.result_columns[self.current_cell]['current_value']
        self.hg_pure_no_list = re.findall("\d+\d{3}\d+", self.unmodified_hg)
        for hg in self.hg_pure_no_list:
            if not self.modified_hg=='':
                self.modified_hg += '/' + hg
            else:
                self.modified_hg = hg
        # while not str(self.unmodified_hg)[0].isnumeric():
        #     self.unmodified_hg=self.unmodified_hg[1:]

        # for i, xchar in enumerate(self.unmodified_hg):
        #     if not xchar.isnumeric():
        #         self.unmodified_hg=self.unmodified_hg[:i]

        # Write it to result sheet
        # self.result_columns[self.current_cell]['current_value']=self.unmodified_hg
        if self.is_win32==False:
            self.result_sh.cell(
                row=self.result_next_row,
                column=self.result_columns[self.current_cell]['colno']-1,
                value=self.modified_hg)
            
            self.to_transfer[self.current_excel_file_name][kwargs['order_no']][self.result_columns[self.current_cell]['colno']-1]=self.modified_hg
        else:
            self.material_table.Cells(self.result_next_row, self.result_columns[self.current_cell]['colno']-1).Value=self.modified_hg

        # print("Inside hg_works")
        pass

    def imp_bp_works(self, **kwargs):
        self.unmodified_imp_bp=self.result_columns[self.current_cell]['current_value']
        if str(self.unmodified_imp_bp).lower().__contains__('imp'):
            self.modified_imp_bp='IMP'
        else:
            self.modified_imp_bp='BP'

        self.result_columns[self.current_cell]['current_value']=self.modified_imp_bp
        # print("Inside imp_bp works")

    def is_rows_queue(self, number_list):
        """
            rows_list: it must be list of numbers only
        """
        is_rows_queue=True
        number_list.sort()
        old_number=0
        for number in number_list:
            difference=number-old_number
            
            if not old_number==0 and not difference==1:
                is_rows_queue=False
                break
            
            old_number=number

        if is_rows_queue:
            return True
        else:
            return False

    def is_ordernos_queue(self, number_list):
        """
            order_numbers_list: it must be list of numbers only
        """
        is_ordernes_queue=True
        number_list.sort()

        # if last item is not equal to len (only in order numbers)
        if not number_list[0]==1:
            return False
        elif not number_list[len(number_list)-1]==len(number_list):
            return False

        old_number=0
        for number in number_list:
            difference=number-old_number
            
            if not old_number==0 and not difference==1:
                is_ordernes_queue=False
                break
            
            old_number=number

        if is_ordernes_queue:
            return True
        else:
            return False

    def test_transfer_dictionary(self):
        for facture in self.to_transfer:
            for order in self.to_transfer[facture]:
                for column in self.to_transfer[facture][order]:
                    # for data in self.to_transfer[facture][order][column]:
                    print(facture, ' - ', order, ' - ', column, ' - ', self.to_transfer[facture][order][column])
                    # print(self.to_transfer[facture][order][column])
                    pass

if __name__ == '__main__':
    start=time.perf_counter()

    testpath = Path('D:\\BYTK_Facturation\\7. MT\\xxx')
    all_factures_path=Path('D:\\BYTK_Facturation\\7. MT\\1-FACTURE')
    hundreds_file_path = Path('D:\\BYTK_Facturation\\7. MT\\yyy')
    material_table_path=Path.cwd() / 'MATERIAL TABLE_NCC - PQ - test.xlsm'
    hungry_man = InfoFetcher(material_table=str(material_table_path), all_factures_path=testpath)
    hungry_man.create_result_wb_for_result()

    for excel in hungry_man.list_combined_path_of_factures_to_transfer:
        if (excel.suffix).lower == 'xls':
            print('Skipped xls file: ', excel)
            continue
        try:
            testanalyze = AnalyzeForDataFrame(
                type='df', wb=pd.ExcelFile(excel), extension=excel.suffix)
        except:
            print('testanalyze failed: ', excel)
            pass
        if testanalyze.is_facture:
            print("Started memorizing: ", excel.name)
            hungry_man.eat_facture(df_facture=testanalyze.get_facture_df(), df_annexe=testanalyze.get_annexe_df(), current_excel_file_name=excel.name)

    hungry_man.save_result_wb_after_done()
    hungry_man.transfer_to_material_table()
    finish=time.perf_counter()
    print('Total time spent {} seconds'.format(round(finish-start, 0)))
