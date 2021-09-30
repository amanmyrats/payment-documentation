# import win32com.client as win32
from pathlib import Path
import os
import re
import openpyxl
from openpyxl import load_workbook, Workbook
import xlwings as xw
import time
import pandas as pd
import numpy as np
import img2pdf
from fpdf import FPDF

boq = '03.503'
print(boq)
print(boq.replace(' ', ''))
print(boq.replace('\s', '').split(';'))

modified_boq=''
for b in boq.split(';'):
    modified_boq += b + '\n'

print('this is mofified:',modified_boq)
# boq = 'new'
# # save FPDF() class into a variable pdf
# pdf = FPDF()
# # Add a page
# pdf.add_page()
# # set style and size of font that you want in the pdf
# pdf.set_font("Arial", "B", size = 60)
# # create a cell
# pdf.cell(200, 250, txt = boq, ln = 2, align = 'C')
# pdf.output(boq + ".pdf")   

# xpath = Path(r'G:\7. MT\5-CO')

# for i, image in enumerate(xpath.rglob('*.*'), start=1):
#     if 'jpg' in str(image.suffix).lower():
#         print('JPG : ', i)
#         pdf_name = str(image.stem) + ('.pdf')
#         result_pdf_full_path = image.parent / pdf_name
#         if not Path(result_pdf_full_path).exists():
#             print('Working with: ', image.name)
#             with open(result_pdf_full_path, 'wb') as pdf:
#                 pdf.write(img2pdf.convert(str(image)))




# xl = Path(r'D:\BYTK_Facturation\7. MT\3-DECLARATION\07212-041120-0014918.PDF')

# print(xl.name)
# print(xl.stem)
# print(xl.suffix)
# print(f'{xl.stem}{str(xl.suffix).lower()}')
# print(f'{xl.stem}.{xl.suffix}')

# rs = pd.DataFrame()

# if rs['path']:
#     print('it does not raise error')
# # print(rs['path'])

# s = '001-TK-NCC '
# print(f'test{s}test'.ljust(10))
# print(f'test{s[:-1]}test'.ljust(10))

# t = Workbook()
# sh = t.active

# sh.cell(row=1, column=1).value = 'aman'
# t.save('test workbook.xlsx')


# def decl_splitter( *args):
#     decl_pattern = r'\d+[/-]\d+[/-]\d+'
#     return re.findall(decl_pattern, args[0])

# no = '07212/030820/0009315 (part)07212/070920/0011281 (part)07212/190920/0012007 (уничтожение)'
# no1 = '07212/030719/0009061'
# no2 = '"IMP-TEMP-07212/021018/0008850IMP-DEF-07212/210420/0005588"'
# no3 = '07212/030221/0001840 (part)\n07212/170221/0002745 (part)'

# print(decl_splitter(no))
# print(decl_splitter(no1))
# print(decl_splitter(no2))
# print(decl_splitter(no3))

# def tds_splitter( *args):
#     tds_pattern = r'\d+\.\d+\.\d+\.\d+'
#     return re.findall(tds_pattern, args[0])

# no = '795.5.134.12317 (part)795.5.134.11792 (part)'
# no1 = '795.5.134.11278 (part)\n795.5.134.12491 (part)\n795.5.154.13243 (part)'
# no2 = '795.5.154.03449 (part)LETTRE 08/795'
# print(tds_splitter(no))
# print(tds_splitter(no1))
# print(tds_splitter(no2))




# def routage_splitter( *args):
#     check_slash_pattern = r'[/-]\d{2}[a-zA-Z]{2,3}'
#     slash_as_order_seperator_pattern = r'\d{2}[a-zA-z]{2,3}\d{2,3}(?:[/-]{0,1}\d{1,2}){0,1}'
#     slash_as_routage_serepator_pattern = r'\d{2}[a-zA-z]{2,3}\d{2,3}'

#     if len(re.findall(check_slash_pattern, args[0]))>0:
#         # Then the slash is only routage seperator,
#         return re.findall(slash_as_routage_serepator_pattern, args[0])
#     else:
#         # Slash is order seperator
#         return re.findall(slash_as_order_seperator_pattern, args[0])

# no = '21ROA133'
# no1 = '20ROA130\n20ROA131/1\n20ROA132/1\n20ROA133/1\n20ROA134/1'
# no2 = '20BAA054/20BAA055'
# print(routage_splitter(no))
# print(routage_splitter(no1))
# print(routage_splitter(no2))




# import psutil
# from helper import available_ram
# from pdfrw import PdfReader, PdfWriter
# import gc
# import tempfile
# import PyPDF2
# import cv2

# img = cv2.imread()


# path1 = open(r'D:\BYTK_Facturation\to\to Enesh\ANNEXE 2 - RPM - Rapport Photo n45  du 26.02.2021_FR+RU+TKM.docx.pdf', 'rb')
# # path2 = r'D:\BYTK_Facturation\to\to Enesh\ANNEXE 2 - RPM - Rapport Photo n45  du 26.02.2021_FR+RU+TKM resized.docx.pdf'

# from io import BytesIO

# tmp = BytesIO()

# merger = PyPDF2.PdfFileMerger()
# merger.append(fileobj=path1)
# # merger.append(fileobj=path)
# merger.write(tmp)

# PyPDF2.filters.compress(tmp.getvalue())
# merger.write(open(r"D:\BYTK_Facturation\to\to Enesh\test_out2.pdf", 'wb'))
# # merger = PyPDF2.PdfFileReader()
# PyPDF2.filters.compress(tmp.getvalue())




# print(tempfile.gettempdir())
# print(type(tempfile.gettempdir()))
# print(type(Path(tempfile.gettempdir()) / 'gen_py'))
# path = Path(r'D:\BYTK_Facturation\7. MT\payment')
# writer = PdfWriter()
# for pdf in path.glob('*.pdf'):
#     reader=PdfReader(pdf)
#     print(reader.Size)
    
#     gc.collect()
#     writer.addpages(reader.pages)
#     reader=None
#     gc.collect()
    # print('writer size: ', writer.killobj['Size'])

# writer.write('testtets.pdf')

# print(available_ram(type='percentage'))


# print(psutil.virtual_memory()._asdict()['available']/1024/1024/1024)
# mem=str(os.popen('free -t -m').readlines())
# xfile=Path(r'D:\BYTK_Facturation\7. MT\Concrete-Unit Price Calculation.xlsx')
# def add_prename_to_full_file_name( full_path, prename):
#         parent=full_path.parent
#         name=full_path.name
#         new_name='{prename}{name}'.format(prename = prename, name = name)
#         prename_added_full_path = parent / new_name
#         return prename_added_full_path

# print(add_prename_to_full_file_name(full_path=xfile, prename='1-'))
# print(add_prename_to_full_file_name(full_path=xfile, prename='1-').name)

# def find_decl_page( haryt_no):
#         if haryt_no==1:
#             return 1
#         else:
#             temp_mod = (haryt_no-1) % 3
#             temp_divided = int((haryt_no-1) / 3)
#             if temp_mod==0:
#                 return temp_divided+1
#             else:
#                 return temp_divided+2

# for i in range(1,30):
#     print(find_decl_page(i))

# xfile=Path(r'D:\BYTK_Facturation\7. MT\Concrete-Unit Price Calculation.xlsx')

# stm=xfile.stem 
# sfx=xfile.suffix 

# print(stm)
# print(sfx)


# try:
#     from PIL import Image
# except ImportError:
#     import Image
# import pytesseract

# pytesseract.pytesseract.tesseract_cmd = r'<full_path_to_your_tesseract_executable>'


# import pytesseract
# from pdf2image import convert_from_path
# import glob

# pdfs = glob.glob(r"D:\BYTK_Facturation\7. MT\all facture names.pdf")

# for pdf_path in pdfs:
#     pages = convert_from_path(pdf_path, 500)

#     for pageNum,imgBlob in enumerate(pages):
#         text = pytesseract.image_to_string(imgBlob,lang='eng')

#         with open(f'{pdf_path[:-4]}_page{pageNum}.txt', 'w') as the_file:
#             the_file.write(text)



# from pdfminer.high_level import extract_text

# text = extract_text(r'D:\BYTK_Facturation\7. MT\all facture names.pdf')

# print(text)





# from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
# from pdfminer.converter import TextConverter
# from pdfminer.layout import LAParams
# from pdfminer.pdfpage import PDFPage
# from io import StringIO

# def convert_pdf_to_txt(path):
#     rsrcmgr = PDFResourceManager()
#     retstr = StringIO()
#     codec = 'utf-8'
#     laparams = LAParams()
#     device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
#     fp = open(path, 'rb')
#     interpreter = PDFPageInterpreter(rsrcmgr, device)
#     password = ""
#     maxpages = 0
#     caching = True
#     pagenos=set()

#     for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password,caching=caching, check_extractable=True):
#         interpreter.process_page(page)

#     text = retstr.getvalue()
#     print(text)

#     fp.close()
#     device.close()
#     retstr.close()
#     return text

# path=r'D:\BYTK_Facturation\7. MT\payment\04.100\NCC-TK-3408\07212-110320-0003600.PDF'
# convert_pdf_to_txt(path)



# xlist=[1,3,22,4]

# print(xlist.index(22))

# print(np.argmax(xlist))


# start=time.perf_counter()

# xpath=Path(r'D:\BYTK_Facturation\7. MT\1-FACTURE')

# xlist=[x for x in xpath.rglob('*.*')]

# # print(xlist)
# print('length: ', len(xlist))
# # for x, xxx in enumerate(xlist, start=1):
# #     print(x, ' - ', xxx)

# df=pd.DataFrame(xlist, columns=['facture_path'])
# print('Daraframe\'s shape: ',df.shape)
# print(df.head())
# finish=time.perf_counter()

# print('Time spent: ', finish-start)

# xx = 'D:\\BYTK_Facturation\\7. MT\\GENERAL INVOICE rev19.xlsx'

# wb = load_workbook(Path(xx))
# xset=set()

# xset.add('aman')
# xset.add('myrat')
# xset.add('aman')

# print(xset)
# list1=['aman', 'myrat', 'ahmet', 'aman']

# list2=['aman']

# list1.remove('aman')

# print(list1)

# # xstr='МРАМОРBOTTICINOPOLISHEDSMSMSMSMSM'
# xstr='МРАМОР BIANCO CARRARA 64,6*92,3*3SM, 69,4*48,6*3SM, 49,1*49,1*3SM, 52,4*49,1*3SM, 50,9*49,1*3SM '
# ral="АаБбВвГгДдЕеЁёЖжЗзИиЙйКкЛлМмНнОоПпСсТтУуФфХхЦцЧчШшЩщЪъЫыЬьЭэЮюЯя"
# if bool(re.search('[а-яА-Я]', xstr)):
#     print('has russian')
# print('length: ', len(xstr))
# xxx=0
# for x in xstr:
#     if x in ral:
#         xxx+=1
#         print('russian: ', x)

# print(int(xxx/len(xstr)*100))


# xl=Path.cwd() / 'MATERIAL TABLE_NCC - PQ - test.xlsm'

# excel=load_workbook(xl)

# # self.sheet.range(self.temp_row, self.temp_col).value=self.temp_val

# print(excel.active)

# dict={}

# try:
#     dict['k1']['k2']['k3']='test3'
# except:
#     try:
#         dict['k1']={'k2':{}}



# print(dict)


# number='3.1aer2'

# print(type(float(number)))

# sourcefile=Path.cwd() / 'xxx/1481-TK-NCC TK modife 1.xlsm'
# destfolder=Path.cwd() / 'xxx/dest'


# xl=win32.gencache.EnsureDispatch('Excel.Application')
# # xl.Visible=False
# xl.Visible=True
# wb=xl.Workbooks.Open(sourcefile)

# dest_pdf_name=str(sourcefile.stem) + '.pdf'

# xlTypePDF = 0
# xlQualityStandard = 0

# dest_full_name= r'' + str(destfolder / dest_pdf_name) + ''

# print('dest_full_name: ', dest_full_name)
# # wb.ActiveSheet.ExportAsFixedFormat(0, dest_full_name)
# sh_facture=wb.Sheets('facture')

# print(sh_facture.Cells(8,3).Value)
# x=destfolder / dest_pdf_name
# print(x)
# # os.system('pause')

# wb.Saved=True
# # wb.Close()

# # xl.Quit()
