import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import pymsgbox
import sys
import re


class AllFiles:
    def __init__(self, *args, **kwargs):

        # Assign all pdf files into a dictionary
        self.ginv = {}
        self.wb = kwargs['wb']
        self.mt_folder = Path(kwargs['mt_folder'])
        

        # self.wb = r'D:\BYTK_Facturation\7. MT\GENERAL INVOICE rev19.xlsx'
        # self.mt_folder = Path(r'D:\BYTK_Facturation\7. MT')
        
        self.is_wb_error=False
        self.all_folders_in_place=True
        self.partially_found=False
        self.none_found=False

        # Prefix numbers for destination files
        self.prefix_dict = {'facture': '1-', 'routage': '2-',
                            'decl': '3-', 'tds': '4-', 'coo': '5-', }

        # Assign workbook, if it is not a workbook then warn user
        try:
            self.do_wb_works()
        except:
            self.is_wb_error=True

        # Assign paths to dictionary, and check if it is a right dictionary, if it is not right then warn user.
        self.do_path_works()
        if self.path_not_found:
            self.all_folders_in_place=False
            if len(self.found_paths) > 0:
                self.partially_found=True
                self.partially_found_alert_message="Inside your folder there is no {}".format(list(map(lambda x: x, self.not_found_paths))) 
                self.partially_found_alert_message+='\n'
                self.partially_found_alert_message+="But there is folder of {}".format(list(map(lambda x: x, self.found_paths)))
            else:
                self.none_found=True
                self.none_found_alert_message="Inside your folder there is no {}".format(
                    list(map(lambda x: x, self.not_found_paths)))

    def do_wb_works(self):

        self.wb = load_workbook(filename=self.wb, read_only=False, keep_vba=True)

        self.ws = self.wb['GENERAL INVOICE']
        self.want_replace = False

        self.not_found_files_dict = {'facture': [],
                                'routage': [],
                                'decl' : [],
                                'tds' : [],
                                'coo' : []}

        self.total_factures = len(self.ws['A'])

        # re.findall("\d+\d{3}\d+", self.unmodified_hg)

        for i in range(3, self.total_factures+1):
            print(self.ws.cell(row=i, column=1).value)
            # root
            self.root_facture = self.remove_slash_and_newline(
                self.ws.cell(row=i, column=1).value)
            # facture
            self.facture_no=self.root_facture

            # routage
            self.routage_no = self.remove_slash_and_newline(
                self.ws.cell(row=i, column=2).value)
            
            # decl
            self.decl_no = self.remove_slash_and_newline(
                self.ws.cell(row=i, column=3).value)
            
            # tds
            self.tds_no = self.remove_slash_and_newline(
                self.ws.cell(row=i, column=4).value)
            
            # coo
            self.coo_no = self.remove_slash_and_newline(
                self.ws.cell(row=i, column=5).value)

            self.ginv[self.root_facture] = {'facture': {'no': self.facture_no}}
            self.ginv[self.root_facture].update(
                {'routage': {'no': self.routage_no}})
            self.ginv[self.root_facture].update({'decl': {'no': self.decl_no}})
            self.ginv[self.root_facture].update({'tds': {'no': self.tds_no}})
            self.ginv[self.root_facture].update({'coo': {'no': self.coo_no}})

    # Row and column numbers
            # ROW
            try:
                self.ginv[self.root_facture]['row'] = i
            except:
                print('Could not write row number of root facture {} into ginv dictionary, in FilesClass.py'.format(self.root_facture))
                pass
            
        # COLUMN
        # Decide which column should I use when writing 'NOT FOUND' note
        self.facture_not_found_column = 24
        self.routage_not_found_column = 25
        self.decl_not_found_column = 26
        self.tds_not_found_column = 27
        self.coo_not_found_column = 28

        self.not_found_columns = {'facture': self.facture_not_found_column,
                                'routage': self.routage_not_found_column,
                                'decl': self.decl_not_found_column,
                                'tds': self.tds_not_found_column,
                                'coo': self.coo_not_found_column}

        self.not_found_column_titles_dict = {'facture': 'Facture Status', 
                                            'routage': 'Routage Status',
                                            'decl': 'Declaration Status',
                                            'tds': 'TDS Status',
                                            'coo': 'COO Status'}
        for type in self.not_found_column_titles_dict:
            self.temp_found = False
            for xr in range(1, self.ws.max_row + 1):
                if not self.temp_found:
                    for xc in range(1, self.ws.max_column + 1):
                        if self.not_found_column_titles_dict[type] == self.ws.cell(row=xr, column=xc).value:
                            self.temp_found = True
                            self.not_found_columns[type] = xc
                            break
                else:
                    break

    def do_path_works(self):
        # Define source and destination paths for pdf files
        self.overall_path = {}

        self.overall_path = {'source': {'facture': self.mt_folder / '1-FACTURE',
                                        'routage': self.mt_folder / '2-CMR',
                                        'decl': self.mt_folder / '3-DECLARATION',
                                        'tds': self.mt_folder / '4-TDS',
                                        'coo': self.mt_folder / '5-CO'},
                             'destination': {'parent': self.mt_folder / '00_All in One', 'sub': self.mt_folder / ''}
                             }

        # Check if user selected right folder
        self.path_not_found=False
        self.not_found_paths=[]
        self.found_paths=[]
        for key in self.overall_path['source']:
            if not Path(self.overall_path['source'][key]).exists():
                self.path_not_found=True
                self.not_found_paths.append(key)
            else:
                self.found_paths.append(key)
            

    def remove_slash_and_newline(self, *args):
        for xstr in args:
            self.xstr = str(xstr)
            self.xstr = self.xstr.replace('/', '-')
            self.xstr = self.xstr.replace('\n', ' ')
        return self.xstr
    
    def routage_splitter(self, *args):
        # Check the task of slash, does it seperate two routage
        # or is it only seperator one routage.
        check_slash_pattern = r'[/-]\d{2}[a-zA-Z]{2,3}'
        slash_as_order_seperator_pattern = r'\d{2}[a-zA-z]{2,3}\d{2,3}(?:[/-]{0,1}\d{1,2}){0,1}'
        slash_as_routage_serepator_pattern = r'\d{2}[a-zA-z]{2,3}\d{2,3}'

        if len(re.findall(check_slash_pattern, args[0]))>0:
            # Then the slash is only routage seperator,
            # print('As seperator {}'.format(args[0]))
            # print('Seperator result {}'.format(re.findall(slash_as_routage_serepator_pattern, args[0])))
            return re.findall(slash_as_routage_serepator_pattern, args[0])
        else:
            # Slash is order seperator
            # print('As order slash {}'.format(args[0]))
            # print('Order slash result {}'.format(re.findall(slash_as_order_seperator_pattern, args[0])))
            return re.findall(slash_as_order_seperator_pattern, args[0])

    def tds_splitter(self, *args):
        tds_pattern = r'\d+\.\d+\.\d+\.\d+'
        return re.findall(tds_pattern, args[0])

    def decl_splitter(self, *args):
        decl_pattern = r'\d+[/-]\d+[/-]\d+'
        return re.findall(decl_pattern, args[0])
